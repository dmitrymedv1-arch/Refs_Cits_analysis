import streamlit as st
import pandas as pd
import numpy as np
import requests
import json
import time
from typing import List, Dict, Any, Tuple, Set
from datetime import datetime
from collections import Counter
import re
import os
import tempfile
import zipfile
import shutil
import io
import csv
import base64
import logging
from functools import lru_cache
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from habanero import Crossref
from crossref_commons.retrieval import get_publication_as_json
from ratelimit import limits, sleep_and_retry
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import tqdm

# Ensure NLTK data is available
try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

# =============================================
# CONFIGURATION AND SETUP
# =============================================

class Config:
    REQUEST_TIMEOUT = 30
    MAX_RETRIES = 5
    DELAY_BETWEEN_REQUESTS = 0.2
    RETRY_DELAY = 1

# =============================================
# CACHE MANAGEMENT
# =============================================

class CacheManager:
    def __init__(self):
        self.cache_dir = tempfile.mkdtemp()
        self.cache_files = {}
    
    def save_cache(self, key, data):
        cache_file = os.path.join(self.cache_dir, f"{key}.pkl")
        try:
            import pickle
            with open(cache_file, 'wb') as f:
                pickle.dump(data, f)
            self.cache_files[key] = cache_file
        except Exception as e:
            st.warning(f"Could not save cache for {key}: {e}")
    
    def load_cache(self, key):
        cache_file = os.path.join(self.cache_dir, f"{key}.pkl")
        if os.path.exists(cache_file):
            try:
                import pickle
                with open(cache_file, 'rb') as f:
                    return pickle.load(f)
            except Exception as e:
                st.warning(f"Could not load cache for {key}: {e}")
        return None
    
    def clear_cache(self):
        try:
            shutil.rmtree(self.cache_dir)
            self.cache_dir = tempfile.mkdtemp()
            self.cache_files = {}
        except Exception as e:
            st.warning(f"Could not clear cache: {e}")

# Initialize cache manager
cache_manager = CacheManager()

# =============================================
# PERFORMANCE MONITORING
# =============================================

class PerformanceMonitor:
    def __init__(self):
        self.start_time = None
        self.request_count = 0

    def start(self):
        self.start_time = datetime.now()

    def increment_request(self):
        self.request_count += 1

    def get_stats(self):
        if self.start_time:
            elapsed = (datetime.now() - self.start_time).total_seconds()
            return {
                'total_requests': self.request_count,
                'elapsed_seconds': elapsed,
                'elapsed_minutes': elapsed / 60,
                'requests_per_second': self.request_count / elapsed if elapsed > 0 else 0
            }
        return {}

# =============================================
# AFFILIATION PROCESSOR (CORRECTED VERSION)
# =============================================

class FastAffiliationProcessor:
    """Fast affiliation processor with proper organization grouping"""

    def __init__(self):
        self.common_keywords = {
            'university', 'college', 'institute', 'school', 'department', 'faculty',
            'laboratory', 'center', 'centre', 'academy', 'universität', 'universitat',
            'université', 'universite', 'polytechnic', 'technical', 'technology',
            'research', 'science', 'sciences', 'studies', 'medical', 'hospital',
            'clinic', 'foundation', 'corporation', 'company', 'inc', 'ltd', 'corp'
        }
        self.organization_cache = {}
        
        # Country mapping from the working code
        self.COUNTRY_MAPPING = {
            'CN': 'China', 'RU': 'Russia', 'IT': 'Italy', 'US': 'USA', 'GB': 'United Kingdom',
            'DE': 'Germany', 'FR': 'France', 'JP': 'Japan', 'CA': 'Canada', 'AU': 'Australia',
            'NL': 'Netherlands', 'ES': 'Spain', 'SE': 'Sweden', 'CH': 'Switzerland', 'KR': 'South Korea',
            'IN': 'India', 'BR': 'Brazil', 'MX': 'Mexico', 'AR': 'Argentina', 'ZA': 'South Africa',
            'PL': 'Poland', 'IL': 'Israel', 'SG': 'Singapore', 'NO': 'Norway', 'DK': 'Denmark',
            'FI': 'Finland', 'BE': 'Belgium', 'AT': 'Austria', 'PT': 'Portugal', 'GR': 'Greece',
            'TR': 'Turkey', 'SA': 'Saudi Arabia', 'AE': 'UAE', 'CL': 'Chile', 'CO': 'Colombia',
            'NZ': 'New Zealand', 'IE': 'Ireland', 'CZ': 'Czech Republic', 'HU': 'Hungary', 'RO': 'Romania',
            'UA': 'Ukraine', 'TH': 'Thailand', 'MY': 'Malaysia', 'ID': 'Indonesia', 'PH': 'Philippines',
            'VN': 'Vietnam', 'EG': 'Egypt', 'MA': 'Morocco', 'KE': 'Kenya', 'NG': 'Nigeria',
        }

    def get_country_display(self, code):
        """Get country display name from code"""
        if not code or code == 'Unknown':
            return 'Unknown'
        code = code.upper()
        return self.COUNTRY_MAPPING.get(code, code)

    def extract_main_organization_fast(self, affiliation: str) -> str:
        """Fast extraction of main organization from full affiliation"""
        if not affiliation or affiliation in ['Unknown', 'Error', '']:
            return "Unknown"

        # Cache results
        if affiliation in self.organization_cache:
            return self.organization_cache[affiliation]

        # Clean text
        clean_affiliation = affiliation.strip()

        # Remove email addresses
        clean_affiliation = re.sub(r'\S+@\S+', '', clean_affiliation)

        # Remove postal codes and addresses
        clean_affiliation = re.sub(r'\d{5,}(?:-\d{4})?', '', clean_affiliation)
        clean_affiliation = re.sub(r'p\.?o\.? box \d+', '', clean_affiliation, flags=re.IGNORECASE)
        clean_affiliation = re.sub(r'\b\d+\s+[a-zA-Z]+\s+[a-zA-Z]+\b', '', clean_affiliation)

        # Split by commas, semicolons and other separators
        parts = re.split(r'[,;]', clean_affiliation)

        # Look for part with main organization
        main_org_candidates = []

        for part in parts:
            part = part.strip()
            if not part or len(part) < 5:  # Skip too short parts
                continue

            part_lower = part.lower()

            # Check if part contains organization keywords
            has_org_keyword = any(keyword in part_lower for keyword in self.common_keywords)

            if has_org_keyword:
                main_org_candidates.append(part)

        # Choose best candidate
        if main_org_candidates:
            # Prefer longer names (usually full organization name)
            main_org_candidates.sort(key=len, reverse=True)
            main_org = main_org_candidates[0]
        else:
            # If not found by keywords, take first meaningful part
            for part in parts:
                part = part.strip()
                if len(part) > 10:
                    main_org = part
                    break
            else:
                # If still not found, take first non-empty part
                for part in parts:
                    part = part.strip()
                    if part:
                        main_org = part
                        break
                else:
                    main_org = clean_affiliation

        # Clean result
        main_org = re.sub(r'\s+', ' ', main_org).strip()
        main_org = re.sub(r'^[^a-zA-Z0-9]*|[^a-zA-Z0-9]*$', '', main_org)

        result = main_org if main_org else "Unknown"
        self.organization_cache[affiliation] = result
        return result

    def normalize_organization_name(self, org_name: str) -> str:
        """Normalizes organization name for grouping"""
        if not org_name or org_name == "Unknown":
            return org_name

        # Convert to lowercase
        normalized = org_name.lower()

        # Remove common prefixes and suffixes
        remove_patterns = [
            r'^the\s+', r'\s+the$',
            r'\bdept\.?\s+of\b', r'\bdepartment\s+of\b',
            r'\bfaculty\s+of\b', r'\bschool\s+of\b',
            r'\binstitute\s+of\b', r'\binstitution\s+of\b',
            r'\bcollege\s+of\b', r'\bacademy\s+of\b',
            r'\blaboratory\b', r'\blab\b',
            r'\bcenter\b', r'\bcentre\b',
            r'\bdivision\b', r'\bgroup\b',
            r'\binc\.?$', r'\bltd\.?$', r'\bcorp\.?$', r'\bco\.?$',
            r'\bllc\.?$', r'\bgmbh\.?$'
        ]

        for pattern in remove_patterns:
            normalized = re.sub(pattern, '', normalized)

        # Remove extra spaces and characters
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        normalized = re.sub(r'[^\w\s&]', '', normalized)

        return normalized.strip()

    def group_similar_organizations(self, organizations: List[str]) -> Dict[str, List[str]]:
        """Groups similar organizations"""
        if not organizations:
            return {}

        # Create normalized versions
        normalized_map = {}
        for org in organizations:
            if org != "Unknown":
                normalized = self.normalize_organization_name(org)
                if normalized:
                    if normalized not in normalized_map:
                        normalized_map[normalized] = []
                    normalized_map[normalized].append(org)

        # Merge very similar groups
        final_groups = {}
        normalized_keys = list(normalized_map.keys())

        for i, key1 in enumerate(normalized_keys):
            if key1 not in final_groups:
                final_groups[key1] = []

            # Add all organizations for this group
            final_groups[key1].extend(normalized_map[key1])

            # Find similar groups to merge
            for j, key2 in enumerate(normalized_keys[i+1:], i+1):
                if self.are_organizations_similar(key1, key2):
                    if key2 in normalized_map:
                        final_groups[key1].extend(normalized_map[key2])
                    # Mark for deletion
                    if key2 in final_groups:
                        del final_groups[key2]

        return final_groups

    def are_organizations_similar(self, org1: str, org2: str) -> bool:
        """Checks if two organizations are similar"""
        if not org1 or not org2:
            return False

        org1_lower = org1.lower()
        org2_lower = org2.lower()

        # Simple inclusion check
        if org1_lower in org2_lower or org2_lower in org1_lower:
            return True

        # Split into words and check intersection
        words1 = set(org1_lower.split())
        words2 = set(org2_lower.split())

        if not words1 or not words2:
            return False

        # Remove stop words
        stop_words = {'the', 'and', 'of', 'for', 'in', 'on', 'at', 'to', 'by'}
        words1 = words1 - stop_words
        words2 = words2 - stop_words

        if not words1 or not words2:
            return False

        # Calculate Jaccard coefficient
        intersection = words1.intersection(words2)
        union = words1.union(words2)

        similarity = len(intersection) / len(union) if union else 0

        return similarity > 0.6  # Similarity threshold

    def process_affiliations_list_fast(self, affiliations: List[str]) -> Tuple[Dict[str, int], Dict[str, List[str]]]:
        """Fast processing of affiliation list with grouping"""
        if not affiliations:
            return {}, {}

        # Extract main organizations
        main_organizations = []
        for aff in affiliations:
            if aff and aff not in ['Unknown', 'Error']:
                main_org = self.extract_main_organization_fast(aff)
                if main_org and main_org != "Unknown":
                    main_organizations.append(main_org)

        if not main_organizations:
            return {}, {}

        # Group similar organizations
        grouped_organizations = self.group_similar_organizations(main_organizations)

        # Choose representative for each group (most frequent or longest name)
        group_representatives = {}
        for normalized_name, org_list in grouped_organizations.items():
            if org_list:
                # Choose longest name as representative (usually full organization name)
                representative = max(org_list, key=len)
                group_representatives[representative] = org_list

        # Count frequencies for representatives
        frequency_count = {}
        for representative, org_list in group_representatives.items():
            frequency_count[representative] = len(org_list)

        return frequency_count, group_representatives

    def get_affiliations_from_openalex(self, doi: str) -> List[Tuple[str, str]]:
        """Get affiliations from OpenAlex API (from working code)"""
        url = f"https://api.openalex.org/works/doi:{doi}"
        try:
            response = requests.get(url, timeout=15)
            if response.status_code != 200:
                return []
            
            data = response.json()
            unique_affs = set()

            if 'authorships' in data:
                for auth in data['authorships']:
                    for inst in auth.get('institutions', []):
                        name = inst.get('display_name')
                        if not name:
                            ror = inst.get('ror', '')
                            name = ror.split('/')[-1] if '/' in ror else 'Unknown'
                        country = inst.get('country_code', 'Unknown')
                        unique_affs.add((name, country))
            
            return [(name, country) for name, country in unique_affs]
        
        except Exception as e:
            return []

# =============================================
# MAIN CITATION ANALYZER
# =============================================

class CitationAnalyzer:
    def __init__(self, rate_limit_calls=10, rate_limit_period=1):
        self.crossref_cache = {}
        self.openalex_cache = {}
        self.rate_limit_calls = rate_limit_calls
        self.rate_limit_period = rate_limit_period
        self.performance_monitor = PerformanceMonitor()
        self._unique_references_cache = {}
        self._unique_citations_cache = {}
        self.unique_ref_data_cache = {}
        self.unique_citation_data_cache = {}
        self.ltwa_map = None
        self.stop_words = set(stopwords.words('english'))
        self.stemmer = PorterStemmer()
        self.fast_affiliation_processor = FastAffiliationProcessor()
        self.scientific_stopwords = {
            'using', 'based', 'study', 'studies', 'research', 'analysis',
            'effect', 'effects', 'properties', 'property', 'development',
            'application', 'applications', 'method', 'methods', 'approach',
            'review', 'investigation', 'characterization', 'evaluation',
            'performance', 'behavior', 'structure', 'synthesis', 'design',
            'fabrication', 'preparation', 'processing', 'measurement',
            'model', 'models', 'system', 'systems', 'technology', 'material',
            'materials', 'sample', 'samples', 'device', 'devices', 'film',
            'films', 'layer', 'layers', 'surface', 'surfaces', 'interface',
            'interfaces', 'nanoparticle', 'nanoparticles', 'nanostructure',
            'nanostructures', 'composite', 'composites', 'coating', 'coatings'
        }
        self.scientific_stopwords_stemmed = {self.stemmer.stem(word) for word in self.scientific_stopwords}
        self.setup_logging()

    def setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'doi_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def validate_doi(self, doi: str) -> bool:
        """Validates DOI with improved processing"""
        if not doi or not isinstance(doi, str):
            return False

        doi = self.normalize_doi(doi)

        doi_pattern = r'^10\.\d{4,9}/[-._;()/:a-zA-Z0-9]+$'

        if not bool(re.match(doi_pattern, doi, re.IGNORECASE)):
            return False

        if len(doi) < 10:
            return False

        if re.search(r'[^\w\.\-_;()/:]', doi):
            return False

        return True

    def normalize_doi(self, doi: str) -> str:
        """Normalizes DOI, removing prefixes and extra characters"""
        if not doi or not isinstance(doi, str):
            return ""

        doi = doi.strip()

        prefixes = [
            'https://doi.org/', 'http://doi.org/', 'doi.org/',
            'doi:', 'DOI:', 'https://dx.doi.org/', 'http://dx.doi.org/',
        ]

        for prefix in prefixes:
            if doi.lower().startswith(prefix.lower()):
                doi = doi[len(prefix):]
                break

        doi = doi.split('?')[0].split('#')[0]
        doi = doi.strip()

        return doi.lower()

    def parse_doi_input(self, input_text: str, max_dois: int = 200) -> List[str]:
        """Parses DOI input with improved processing of various formats"""
        if not input_text or not isinstance(input_text, str):
            st.error("Error: Input is empty or not a string")
            return []

        lines = input_text.strip().split('\n')
        dois = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            line = line.rstrip('.,;')
            doi_pattern = r'10\.\d{4,9}/[-._;()/:a-zA-Z0-9]+'
            found_dois = re.findall(doi_pattern, line, re.IGNORECASE)

            if found_dois:
                dois.extend(found_dois)
            else:
                if 'doi.org/' in line.lower():
                    doi_part = line.lower().split('doi.org/')[-1]
                    doi_part = doi_part.split('?')[0].split('#')[0].strip()
                    if self.validate_doi(doi_part):
                        dois.append(doi_part)
                elif line.lower().startswith('doi:'):
                    doi_part = line[4:].strip()
                    if self.validate_doi(doi_part):
                        dois.append(doi_part)
                elif self.validate_doi(line):
                    dois.append(line)

        cleaned_dois = []
        for doi in dois:
            normalized_doi = self.normalize_doi(doi)
            if self.validate_doi(normalized_doi):
                cleaned_dois.append(normalized_doi)

        unique_dois = []
        seen = set()
        for doi in cleaned_dois:
            if doi not in seen:
                seen.add(doi)
                unique_dois.append(doi)

        unique_dois = unique_dois[:max_dois]

        if not unique_dois:
            st.error("Error: No valid DOIs found in the input.")
            st.info("Please check the format. Valid examples:")
            st.info("  - 10.1234/abcd.1234")
            st.info("  - https://doi.org/10.1234/abcd.1234")
            st.info("  - doi:10.1234/abcd.1234")
        else:
            st.success(f"Found {len(unique_dois)} valid DOI(s)")
            if len(cleaned_dois) > len(unique_dois):
                st.info(f"Removed {len(cleaned_dois) - len(unique_dois)} duplicate DOI(s)")
            if len(cleaned_dois) > max_dois:
                st.info(f"Limited to first {max_dois} unique DOI(s) from {len(cleaned_dois)} found")

        return unique_dois

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10),
           retry=retry_if_exception_type((requests.exceptions.RequestException,
                                        requests.exceptions.Timeout,
                                        requests.exceptions.ConnectionError)))
    @sleep_and_retry
    @limits(calls=15, period=1)
    def get_openalex_data(self, doi: str) -> Dict:
        """Gets data from OpenAlex with retry attempts"""
        if doi in self.openalex_cache:
            return self.openalex_cache[doi]
        try:
            openalex_url = f"https://api.openalex.org/works/https://doi.org/{doi}"
            response = requests.get(openalex_url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()
            if response.status_code == 404:
                self.openalex_cache[doi] = {}
                return {}
            response.raise_for_status()
            result = response.json()
            self.openalex_cache[doi] = result
            return result
        except requests.exceptions.RequestException as e:
            self.logger.warning(f"OpenAlex request failed for {doi}: {e}")
            if doi not in self.openalex_cache:
                self.openalex_cache[doi] = {}
            return {}
        except Exception as e:
            self.logger.error(f"Unexpected error in get_openalex_data for {doi}: {e}")
            if doi not in self.openalex_cache:
                self.openalex_cache[doi] = {}
            return {}

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10),
           retry=retry_if_exception_type((requests.exceptions.RequestException,
                                        Exception)))
    @sleep_and_retry
    @limits(calls=15, period=1)
    def get_crossref_data(self, doi: str) -> Dict:
        """Gets data from Crossref with retry attempts and improved affiliation processing"""
        if doi in self.crossref_cache:
            return self.crossref_cache[doi]
        try:
            cr = Crossref()
            result = cr.works(ids=doi)
            self.performance_monitor.increment_request()
            data = result['message']

            year = None
            for key in ['published-print', 'published-online', 'issued']:
                if key in data and 'date-parts' in data[key]:
                    date_parts = data[key]['date-parts'][0]
                    year = date_parts[0] if date_parts else None
                    break
            data['publication_year'] = year if year else 'Unknown'

            # Use the corrected affiliation extraction
            affiliations, countries = self.extract_affiliations_from_openalex(doi)
            data['extracted_affiliations'] = affiliations
            data['extracted_countries'] = countries

            self.crossref_cache[doi] = data
            return data
        except Exception as e:
            self.logger.warning(f"Crossref request failed for {doi}: {e}")
            self.crossref_cache[doi] = {'publication_year': 'Unknown', 'extracted_affiliations': [], 'extracted_countries': []}
            return {'publication_year': 'Unknown', 'extracted_affiliations': [], 'extracted_countries': []}

    def extract_affiliations_from_openalex(self, doi: str) -> tuple[List[str], List[str]]:
        """Extracts affiliations and countries from OpenAlex (corrected version)"""
        affiliations = set()
        countries = set()

        try:
            aff_data = self.fast_affiliation_processor.get_affiliations_from_openalex(doi)
            for name, country_code in aff_data:
                if name and name != 'Unknown':
                    # Use the fast processor to extract main organization
                    main_org = self.fast_affiliation_processor.extract_main_organization_fast(name)
                    if main_org and main_org != "Unknown":
                        affiliations.add(main_org)
                
                if country_code and country_code != 'Unknown':
                    # Use short country code instead of full name
                    countries.add(country_code)

        except Exception as e:
            self.logger.debug(f"Error extracting affiliations from OpenAlex: {e}")

        return list(affiliations), list(countries)

    @sleep_and_retry
    @limits(calls=10, period=1)
    def quick_doi_search(self, title: str) -> str:
        """Quick DOI search by title"""
        if not title or title == 'Unknown':
            return None

        url = "https://api.crossref.org/works"
        params = {
            'query.title': title,
            'rows': 1,
            'select': 'DOI,title'
        }
        headers = {'User-Agent': 'CitationAnalyzer/1.0 (mailto:your@email.com)'}

        try:
            response = requests.get(url, params=params, headers=headers, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()
            response.raise_for_status()
            data = response.json()

            if data['message']['items']:
                doi = data['message']['items'][0]['DOI']
                if self.validate_doi(doi):
                    return doi
            return None
        except:
            return None

    def safe_calculate_annual_citation_rate(self, citation_count, publication_year, current_year=None):
        """Safe calculation of annual citation rate with error handling"""
        try:
            if not isinstance(citation_count, (int, float)) or citation_count == 0:
                return 0.0

            years = self.calculate_years_since_publication(publication_year, current_year)

            if years is None or not isinstance(years, (int, float)) or years <= 0:
                return 0.0

            return round(citation_count / years, 2)
        except (TypeError, ZeroDivisionError, ValueError):
            return 0.0

    def calculate_years_since_publication(self, publication_year: Any, current_year: int = None) -> int:
        """Safe calculation of years since publication"""
        try:
            if current_year is None:
                current_year = datetime.now().year

            if publication_year is None or publication_year == 'Unknown':
                return 1

            year_str = str(publication_year).strip()
            if not year_str or year_str == 'Unknown':
                return 1

            year_match = re.search(r'\b(19|20)\d{2}\b', year_str)
            if year_match:
                year = int(year_match.group())
            else:
                year = int(year_str)

            if 1900 < year <= current_year:
                return max(1, current_year - year)
            else:
                return 1
        except (ValueError, TypeError):
            return 1

    def get_combined_article_data(self, doi: str) -> Dict[str, Any]:
        """Get combined data from both Crossref and OpenAlex with improved affiliation processing"""
        try:
            crossref_data = self.get_crossref_data(doi)
            openalex_data = self.get_openalex_data(doi)

            title = 'Unknown'
            if openalex_data and openalex_data.get('title'):
                title = openalex_data['title']
            elif crossref_data.get('title'):
                title_list = crossref_data['title']
                if title_list:
                    title = title_list[0]

            year = 'Unknown'
            publication_year = None

            if openalex_data and openalex_data.get('publication_year'):
                publication_year = openalex_data['publication_year']
                year = str(publication_year)
            elif crossref_data.get('publication_year') != 'Unknown':
                publication_year = crossref_data['publication_year']
                year = str(publication_year)

            authors = []
            authors_surnames = []
            authors_with_initials = []

            if openalex_data:
                for author in openalex_data.get('authorships', []):
                    name = author.get('author', {}).get('display_name', 'Unknown')
                    if name != 'Unknown':
                        authors.append(name)
                        surname_with_initial = self.extract_surname_with_initial(name)
                        authors_surnames.append(surname_with_initial)
                        authors_with_initials.append(surname_with_initial)

            if not authors and crossref_data.get('author'):
                for author in crossref_data['author']:
                    given = author.get('given', '')
                    family = author.get('family', '')
                    if given or family:
                        name = f"{given} {family}".strip()
                        authors.append(name)
                        surname_with_initial = self.extract_surname_with_initial(name)
                        authors_surnames.append(surname_with_initial)
                        authors_with_initials.append(surname_with_initial)

            authors_str = ', '.join(authors) if authors else 'Unknown'
            authors_surnames_str = ', '.join(authors_surnames) if authors_surnames else 'Unknown'
            authors_with_initials_str = ', '.join(authors_with_initials) if authors_with_initials else 'Unknown'
            author_count = len(authors) if authors else 0

            journal_info = self.get_journal_info_from_crossref(doi)

            _, crossref_citations, openalex_citations = self.get_citation_data(doi)

            # Improved affiliation processing using the corrected method
            affiliations, countries = self.get_enhanced_affiliations_and_countries(openalex_data, crossref_data, doi)

            current_year = datetime.now().year
            years_since_pub = self.calculate_years_since_publication(publication_year, current_year)

            return {
                'doi': doi,
                'title': title,
                'year': year,
                'publication_year': publication_year,
                'authors': authors_str,
                'authors_surnames': authors_surnames_str,
                'authors_with_initials': authors_with_initials_str,
                'author_count': author_count,
                'journal_full_name': journal_info['full_name'],
                'journal_abbreviation': journal_info['abbreviation'],
                'publisher': journal_info['publisher'],
                'citation_count_crossref': crossref_citations,
                'citation_count_openalex': openalex_citations,
                'affiliations': '; '.join(affiliations),
                'countries': countries,
                'years_since_publication': years_since_pub
            }
        except Exception as e:
            return {
                'doi': doi,
                'title': 'Error',
                'year': 'Unknown',
                'publication_year': None,
                'authors': 'Error',
                'authors_surnames': 'Error',
                'authors_with_initials': 'Error',
                'author_count': 0,
                'journal_full_name': 'Error',
                'journal_abbreviation': 'Error',
                'publisher': 'Error',
                'citation_count_crossref': 0,
                'citation_count_openalex': 0,
                'affiliations': 'Error',
                'countries': 'Error',
                'years_since_publication': 1,
                'error': str(e)
            }

    def get_enhanced_affiliations_and_countries(self, openalex_data: Dict, crossref_data: Dict, doi: str) -> tuple[List[str], str]:
        """Enhanced affiliation processing with grouping"""
        try:
            # Get affiliations from OpenAlex using the corrected method
            openalex_affiliations, openalex_countries = self.get_affiliations_and_countries_from_openalex_data(openalex_data, doi)
            crossref_affiliations, crossref_countries = self.get_affiliations_and_countries_from_crossref_data(crossref_data)

            # Combine affiliations
            all_affiliations = []
            if openalex_affiliations and openalex_affiliations != ['Unknown']:
                all_affiliations.extend(openalex_affiliations)
            if crossref_affiliations:
                all_affiliations.extend(crossref_affiliations)

            # Process affiliations through fast processor
            if all_affiliations:
                affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)
                # Use group representatives as final affiliations
                final_affiliations = list(affiliation_frequencies.keys())
            else:
                final_affiliations = ['Unknown']

            # Combine countries - use short codes
            all_countries = set()
            if openalex_countries and openalex_countries != 'Unknown':
                countries_list = openalex_countries.split(';')
                all_countries.update([c.strip() for c in countries_list if c.strip()])
            if crossref_countries:
                all_countries.update(crossref_countries)

            final_countries = ';'.join(sorted(all_countries)) if all_countries else 'Unknown'

            return final_affiliations, final_countries

        except Exception as e:
            self.logger.error(f"Error in enhanced affiliations processing: {e}")
            return ['Unknown'], 'Unknown'

    def get_affiliations_and_countries_from_openalex_data(self, openalex_data: Dict, doi: str) -> tuple[List[str], str]:
        """Extracts affiliations and countries from OpenAlex data using corrected method"""
        try:
            if not openalex_data:
                # Fall back to direct OpenAlex API call
                aff_data = self.fast_affiliation_processor.get_affiliations_from_openalex(doi)
                affiliations = set()
                countries = set()
                
                for name, country_code in aff_data:
                    if name and name != 'Unknown':
                        main_org = self.fast_affiliation_processor.extract_main_organization_fast(name)
                        if main_org and main_org != "Unknown":
                            affiliations.add(main_org)
                    
                    if country_code and country_code != 'Unknown':
                        # Use short country code
                        countries.add(country_code)
                
                return list(affiliations) or ['Unknown'], ';'.join(sorted(countries)) if countries else 'Unknown'
            
            affiliations = set()
            countries = set()

            for authorship in openalex_data.get('authorships', []):
                for institution in authorship.get('institutions', []):
                    display_name = institution.get('display_name', '')
                    country_code = institution.get('country_code', '')

                    if display_name:
                        # Normalize name through our processor
                        main_org = self.fast_affiliation_processor.extract_main_organization_fast(display_name)
                        if main_org and main_org != "Unknown":
                            affiliations.add(main_org)

                    if country_code:
                        # Use short country code
                        countries.add(country_code)

            return list(affiliations) or ['Unknown'], ';'.join(sorted(countries)) if countries else 'Unknown'
        except Exception as e:
            self.logger.debug(f"Error extracting from OpenAlex: {e}")
            return ['Unknown'], 'Unknown'

    def get_affiliations_and_countries_from_crossref_data(self, crossref_data: Dict) -> tuple[List[str], List[str]]:
        """Extracts affiliations and countries from Crossref data"""
        try:
            if not crossref_data:
                return [], []

            affiliations = crossref_data.get('extracted_affiliations', [])
            countries = crossref_data.get('extracted_countries', [])

            return affiliations, countries
        except Exception as e:
            self.logger.debug(f"Error extracting from Crossref: {e}")
            return [], []

    def get_citation_data(self, doi: str) -> tuple:
        try:
            crossref_data = self.get_crossref_data(doi)
            crossref_citations = crossref_data.get('is-referenced-by-count', 0)

            openalex_data = self.get_openalex_data(doi)
            openalex_citations = openalex_data.get('cited_by_count', 0)

            return doi, crossref_citations, openalex_citations
        except:
            return doi, 0, 0

    def get_journal_info_from_crossref(self, doi: str) -> Dict[str, Any]:
        try:
            data = self.get_crossref_data(doi)
            container_title = data.get('container-title', [])
            short_container_title = data.get('short-container-title', [])
            full_name = container_title[0] if container_title else (short_container_title[0] if short_container_title else 'Unknown')
            abbreviation = short_container_title[0] if short_container_title else (container_title[0] if container_title else 'Unknown')
            return {
                'full_name': full_name,
                'abbreviation': abbreviation,
                'publisher': data.get('publisher', 'Unknown'),
                'issn': data.get('ISSN', [None])[0]
            }
        except:
            return {
                'full_name': 'Unknown',
                'abbreviation': 'Unknown',
                'publisher': 'Unknown',
                'issn': None
            }

    @lru_cache(maxsize=1000)
    def extract_surname_with_initial(self, author_name: str) -> str:
        if not author_name or author_name in ['Unknown', 'Error']:
            return author_name
        clean_name = re.sub(r'[^\w\s\-\.]', ' ', author_name).strip()
        parts = clean_name.split()
        if not parts:
            return author_name
        surname = parts[-1]
        initial = parts[0][0].upper() if parts[0] else ''
        return f"{surname} {initial}." if initial else surname

    def get_references_from_crossref(self, doi: str) -> List[Dict[str, Any]]:
        try:
            article_data = get_publication_as_json(doi)
            return article_data.get('reference', [])
        except:
            return []

    @retry(stop=stop_after_attempt(Config.MAX_RETRIES),
           wait=wait_exponential(multiplier=1, min=Config.RETRY_DELAY, max=10))
    def get_citing_articles_from_openalex(self, doi: str) -> List[str]:
        """Gets citing works through OpenAlex API"""
        citing_dois = []
        try:
            work_id = doi.replace('/', '%2F')
            url = f"https://api.openalex.org/works/https://doi.org/{work_id}"

            response = requests.get(url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()

            if response.status_code == 200:
                data = response.json()
                cited_by_count = data.get('cited_by_count', 0)
                work_openalex_id = data.get('id', '')

                if cited_by_count > 0:
                    per_page = 200
                    total_pages = (cited_by_count + per_page - 1) // per_page

                    for page in range(1, total_pages + 1):
                        citing_url = f"https://api.openalex.org/works?filter=cites:{work_openalex_id}&per-page={per_page}&page={page}"

                        response = requests.get(citing_url, timeout=Config.REQUEST_TIMEOUT)
                        self.performance_monitor.increment_request()

                        if response.status_code == 200:
                            citing_data = response.json()
                            results = citing_data.get('results', [])
                            results_count = len(results)

                            for work in results:
                                if work.get('doi'):
                                    citing_dois.append(work['doi'])

                            if results_count < per_page:
                                break

                            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

                        else:
                            if response.status_code == 429:
                                time.sleep(3)
                                page -= 1
                            else:
                                break

                        if len(citing_dois) >= cited_by_count:
                            break

            return citing_dois

        except Exception as e:
            return []

    def get_citing_articles_from_crossref(self, doi: str) -> List[str]:
        """Gets citing works through Crossref API"""
        citing_dois = []
        try:
            url = f"https://api.crossref.org/works/{doi}"
            response = requests.get(url, timeout=Config.REQUEST_TIMEOUT)
            self.performance_monitor.increment_request()

            if response.status_code == 200:
                data = response.json()
                if 'message' in data and 'is-referenced-by' in data['message']:
                    references = data['message']['is-referenced-by']
                    for ref in references:
                        if isinstance(ref, dict) and 'DOI' in ref:
                            citing_dois.append(ref['DOI'])

        except Exception as e:
            pass

        return citing_dois

    def find_citing_articles(self, doi_list: List[str]) -> Dict[str, Dict]:
        """Main function for finding citing articles"""
        results = {}

        for i, doi in enumerate(doi_list, 1):
            doi = doi.strip()
            if not doi:
                continue

            openalex_citations = self.get_citing_articles_from_openalex(doi)
            crossref_citations = self.get_citing_articles_from_crossref(doi)

            all_citations = list(set(openalex_citations + crossref_citations))

            results[doi] = {
                'count': len(all_citations),
                'citing_dois': all_citations
            }

            time.sleep(Config.DELAY_BETWEEN_REQUESTS)

        return results

    def process_citing_articles_sequential(self, doi_list: List[str]) -> tuple[pd.DataFrame, pd.DataFrame, Dict[str, Dict], List[str]]:
        """Processes citing articles sequentially with full statistics"""
        self.performance_monitor.start()

        citing_results = self.find_citing_articles(doi_list)

        all_citing_articles_data = []
        citing_articles_details = []
        all_citing_titles = []

        # Collect all connections with source_doi preservation
        all_citing_connections = []
        for source_doi, source_data in citing_results.items():
            for citing_doi in source_data['citing_dois']:
                all_citing_connections.append({
                    'source_doi': source_doi,
                    'citing_doi': citing_doi
                })

        # Cache unique citing article data
        all_citing_dois = set(conn['citing_doi'] for conn in all_citing_connections)
        citing_data_cache = {}

        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, citing_doi in enumerate(all_citing_dois):
            try:
                status_text.text(f"Processing citing article {i+1}/{len(all_citing_dois)}: {citing_doi}")
                article_data = self.get_combined_article_data(citing_doi)
                citing_data_cache[citing_doi] = article_data
                all_citing_titles.append(article_data['title'])
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
                
                # Update progress
                progress_bar.progress((i + 1) / len(all_citing_dois))
            except Exception as e:
                citing_data_cache[citing_doi] = {
                    'title': 'Error', 'authors': 'Error', 'authors_with_initials': 'Error',
                    'author_count': 0, 'year': 'Unknown', 'journal_full_name': 'Error',
                    'journal_abbreviation': 'Error', 'publisher': 'Error',
                    'citation_count_crossref': 0, 'citation_count_openalex': 0,
                    'years_since_publication': 1, 'affiliations': 'Error', 'countries': 'Error'
                }
                all_citing_titles.append('Error')

        status_text.empty()
        progress_bar.empty()

        # Create citing_articles_df with ALL connections
        for connection in all_citing_connections:
            citing_doi = connection['citing_doi']
            source_doi = connection['source_doi']

            article_data = citing_data_cache.get(citing_doi, {})
            citing_row = {
                'source_doi': source_doi,  # CORRECT - preserve source
                'position': 'N/A',
                'doi': citing_doi,
                'title': article_data.get('title', 'Unknown'),
                'authors': article_data.get('authors', 'Unknown'),
                'authors_with_initials': article_data.get('authors_with_initials', 'Unknown'),
                'author_count': article_data.get('author_count', 0),
                'year': article_data.get('year', 'Unknown'),
                'journal_full_name': article_data.get('journal_full_name', 'Unknown'),
                'journal_abbreviation': article_data.get('journal_abbreviation', 'Unknown'),
                'publisher': article_data.get('publisher', 'Unknown'),
                'citation_count_crossref': article_data.get('citation_count_crossref', 0),
                'citation_count_openalex': article_data.get('citation_count_openalex', 0),
                'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                    article_data.get('citation_count_crossref', 0), article_data.get('publication_year')
                ),
                'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                    article_data.get('citation_count_openalex', 0), article_data.get('publication_year')
                ),
                'years_since_publication': article_data.get('years_since_publication', 1),
                'affiliations': article_data.get('affiliations', 'Unknown'),
                'countries': article_data.get('countries', 'Unknown'),
                'error': None
            }
            all_citing_articles_data.append(citing_row)

        # Create DataFrame with all connections
        citing_articles_df = pd.DataFrame(all_citing_articles_data) if all_citing_articles_data else pd.DataFrame()

        # Create details table
        for source_doi, data in citing_results.items():
            for citing_doi in data['citing_dois']:
                article_data = citing_data_cache.get(citing_doi, {})
                citing_articles_details.append({
                    'source_doi': source_doi,
                    'citing_doi': citing_doi,
                    'citing_title': article_data.get('title', 'Unknown'),
                    'citing_authors': article_data.get('authors_with_initials', 'Unknown'),
                    'citing_year': article_data.get('year', 'Unknown'),
                    'citing_journal': article_data.get('journal_abbreviation', 'Unknown'),
                    'citation_count': article_data.get('citation_count_openalex', 0)
                })

        citing_details_df = pd.DataFrame(citing_articles_details) if citing_articles_details else pd.DataFrame()

        return citing_articles_df, citing_details_df, citing_results, all_citing_titles

    def get_unique_citations(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Gets unique citing articles"""
        if citations_df.empty:
            return pd.DataFrame()

        cache_key = id(citations_df)
        if cache_key not in self._unique_citations_cache:
            citations_df['citation_id'] = citations_df['doi'].fillna('') + '|' + citations_df['title'].fillna('')
            unique_df = citations_df.drop_duplicates(subset=['citation_id'], keep='first').drop(columns=['citation_id'])
            self._unique_citations_cache[cache_key] = unique_df
        return self._unique_citations_cache[cache_key]

    def find_duplicate_citations(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Finds duplicate citing articles (articles that cite multiple analyzed works)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            citations_df['citation_id'] = citations_df['doi'].fillna('') + '|' + citations_df['title'].fillna('')

            # Count how many different source articles each citing article cites
            citation_counts = citations_df.groupby('citation_id')['source_doi'].nunique().reset_index()
            duplicate_citation_ids = citation_counts[citation_counts['source_doi'] > 1]['citation_id']

            if duplicate_citation_ids.empty:
                columns = list(citations_df.columns) + ['frequency']
                columns.remove('citation_id')
                return pd.DataFrame(columns=columns)

            # Create mapping citation_id -> frequency (number of cited source articles)
            frequency_map = citation_counts.set_index('citation_id')['source_doi'].to_dict()

            # Select duplicate citations and keep one record per citing article
            duplicates = citations_df[citations_df['citation_id'].isin(duplicate_citation_ids)].copy()
            duplicates = duplicates.drop_duplicates(subset=['citation_id'], keep='first')

            # Filter invalid records
            duplicates = duplicates[~((duplicates['doi'].isna()) & (duplicates['title'] == 'Unknown'))]

            # Add frequency column
            duplicates['frequency'] = duplicates['citation_id'].map(frequency_map)
            duplicates = duplicates.drop(columns=['citation_id'])

            return duplicates.sort_values(['frequency', 'doi'], ascending=[False, True])
        except Exception as e:
            st.error(f"Error in find_duplicate_citations: {e}")
            return pd.DataFrame()

    def analyze_citation_authors_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of author frequency in citing articles"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            authors_total = citations_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_total = authors_total[authors_total.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_total = authors_total.value_counts().reset_index()
            author_freq_total.columns = ['author_with_initial', 'frequency_total']
            author_freq_total['percentage_total'] = round(author_freq_total['frequency_total'] / total_citations * 100, 2)

            authors_unique = unique_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_unique = authors_unique[authors_unique.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_unique = authors_unique.value_counts().reset_index()
            author_freq_unique.columns = ['author_with_initial', 'frequency_unique']
            author_freq_unique['percentage_unique'] = round(author_freq_unique['frequency_unique'] / total_unique * 100, 2)

            author_freq = author_freq_total.merge(author_freq_unique, on='author_with_initial', how='outer').fillna(0)
            return author_freq[['author_with_initial', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_journals_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of journal frequency in citing articles with additional metrics"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            journals_total = citations_df['journal_abbreviation']
            journals_total = journals_total[journals_total.isin(['Unknown', 'Error']) == False]
            journal_freq_total = journals_total.value_counts().reset_index()
            journal_freq_total.columns = ['journal_abbreviation', 'frequency_total']
            journal_freq_total['percentage_total'] = round(journal_freq_total['frequency_total'] / total_citations * 100, 2)

            journals_unique = unique_df['journal_abbreviation']
            journals_unique = journals_unique[journals_unique.isin(['Unknown', 'Error']) == False]
            journal_freq_unique = journals_unique.value_counts().reset_index()
            journal_freq_unique.columns = ['journal_abbreviation', 'frequency_unique']
            journal_freq_unique['percentage_unique'] = round(journal_freq_unique['frequency_unique'] / total_unique * 100, 2)

            journal_freq = journal_freq_total.merge(journal_freq_unique, on='journal_abbreviation', how='outer').fillna(0)

            # Add additional metrics for citations
            journal_citation_metrics = []
            for journal in journal_freq['journal_abbreviation']:
                journal_articles = unique_df[unique_df['journal_abbreviation'] == journal]

                total_crossref_citations = journal_articles['citation_count_crossref'].sum()
                total_openalex_citations = journal_articles['citation_count_openalex'].sum()

                avg_crossref_citations = journal_articles['citation_count_crossref'].mean() if len(journal_articles) > 0 else 0
                avg_openalex_citations = journal_articles['citation_count_openalex'].mean() if len(journal_articles) > 0 else 0

                journal_citation_metrics.append({
                    'journal_abbreviation': journal,
                    'total_crossref_citations': total_crossref_citations,
                    'total_openalex_citations': total_openalex_citations,
                    'avg_crossref_citations': round(avg_crossref_citations, 2),
                    'avg_openalex_citations': round(avg_openalex_citations, 2)
                })

            journal_metrics_df = pd.DataFrame(journal_citation_metrics)
            journal_freq = journal_freq.merge(journal_metrics_df, on='journal_abbreviation', how='left').fillna(0)

            return journal_freq[['journal_abbreviation', 'frequency_total', 'percentage_total',
                               'frequency_unique', 'percentage_unique', 'total_crossref_citations',
                               'total_openalex_citations', 'avg_crossref_citations', 'avg_openalex_citations']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_affiliations_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of affiliation frequency in citing articles with grouping"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            all_affiliations = []
            for affil_string in citations_df['affiliations']:
                if pd.isna(affil_string) or affil_string in ['Unknown', 'Error', '']:
                    continue

                try:
                    affil_list = affil_string.split(';')
                    for affil in affil_list:
                        clean_affil = affil.strip()
                        if clean_affil and clean_affil not in ['Unknown', 'Error']:
                            all_affiliations.append(clean_affil)
                except Exception:
                    pass

            if not all_affiliations:
                return pd.DataFrame(columns=['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])

            affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)

            affil_data = []
            for affil, freq in affiliation_frequencies.items():
                percentage_total = round(freq / total_citations * 100, 2) if total_citations > 0 else 0
                affil_data.append({
                    'affiliation': affil,
                    'frequency_total': freq,
                    'percentage_total': percentage_total
                })

            unique_affiliations = []
            for affil_string in unique_df['affiliations']:
                if affil_string and affil_string not in ['Unknown', 'Error']:
                    affil_list = affil_string.split(';')
                    unique_affiliations.extend([affil.strip() for affil in affil_list if affil.strip()])

            if unique_affiliations:
                unique_frequencies, _ = self.fast_affiliation_processor.process_affiliations_list_fast(unique_affiliations)

                affil_df = pd.DataFrame(affil_data)
                for affil, freq in unique_frequencies.items():
                    percentage_unique = round(freq / total_unique * 100, 2) if total_unique > 0 else 0
                    if affil in affil_df['affiliation'].values:
                        affil_df.loc[affil_df['affiliation'] == affil, 'frequency_unique'] = freq
                        affil_df.loc[affil_df['affiliation'] == affil, 'percentage_unique'] = percentage_unique
                    else:
                        affil_df = pd.concat([affil_df, pd.DataFrame([{
                            'affiliation': affil,
                            'frequency_total': 0,
                            'percentage_total': 0,
                            'frequency_unique': freq,
                            'percentage_unique': percentage_unique
                        }])], ignore_index=True)
            else:
                affil_df = pd.DataFrame(affil_data)
                affil_df['frequency_unique'] = affil_df['frequency_total']
                affil_df['percentage_unique'] = affil_df['percentage_total']

            affil_df = affil_df.fillna(0)
            return affil_df[['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)

        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_countries_frequency(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of country frequency in citing articles with separation into individual countries and collaborations"""
        try:
            if citations_df.empty:
                return pd.DataFrame()

            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)

            # Analysis of individual countries
            single_countries = []
            collaborations = []

            for countries in citations_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries.extend(country_list)
                    else:
                        collaborations.append(countries)

            # Frequency of individual countries
            single_country_counter = Counter(single_countries)
            single_country_freq = pd.DataFrame({
                'countries': list(single_country_counter.keys()),
                'type': ['single'] * len(single_country_counter),
                'frequency_total': list(single_country_counter.values())
            })
            single_country_freq['percentage_total'] = round(single_country_freq['frequency_total'] / total_citations * 100, 2)

            # Frequency of collaborations
            collaboration_counter = Counter(collaborations)
            collaboration_freq = pd.DataFrame({
                'countries': list(collaboration_counter.keys()),
                'type': ['collaboration'] * len(collaboration_counter),
                'frequency_total': list(collaboration_counter.values())
            })
            collaboration_freq['percentage_total'] = round(collaboration_freq['frequency_total'] / total_citations * 100, 2)

            # Combine
            country_freq_total = pd.concat([single_country_freq, collaboration_freq], ignore_index=True)

            # Similarly for unique articles
            single_countries_unique = []
            collaborations_unique = []

            for countries in unique_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries_unique.extend(country_list)
                    else:
                        collaborations_unique.append(countries)

            single_country_counter_unique = Counter(single_countries_unique)
            collaboration_counter_unique = Counter(collaborations_unique)

            single_country_freq_unique = pd.DataFrame({
                'countries': list(single_country_counter_unique.keys()),
                'frequency_unique': list(single_country_counter_unique.values())
            })
            single_country_freq_unique['percentage_unique'] = round(single_country_freq_unique['frequency_unique'] / total_unique * 100, 2)

            collaboration_freq_unique = pd.DataFrame({
                'countries': list(collaboration_counter_unique.keys()),
                'frequency_unique': list(collaboration_counter_unique.values())
            })
            collaboration_freq_unique['percentage_unique'] = round(collaboration_freq_unique['frequency_unique'] / total_unique * 100, 2)

            country_freq_unique = pd.concat([
                single_country_freq_unique.assign(type='single'),
                collaboration_freq_unique.assign(type='collaboration')
            ], ignore_index=True)

            country_freq = country_freq_total.merge(country_freq_unique, on=['countries', 'type'], how='outer').fillna(0)

            # Add full country names
            country_freq['country_fullname'] = country_freq['countries'].apply(
                lambda x: ';'.join([self.fast_affiliation_processor.get_country_display(code) for code in x.split(';')])
            )

            return country_freq[['countries', 'country_fullname', 'type', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_citation_year_distribution(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of year distribution for citing articles (from new to old)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()
    
            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)
            
            # Обработка годов для всех цитирований
            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, 2026)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                year_counts_total = years_total.value_counts().reset_index()
                year_counts_total.columns = ['year', 'frequency_total']
                year_counts_total['percentage_total'] = round(year_counts_total['frequency_total'] / total_citations * 100, 2)
            else:
                year_counts_total = pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total'])
    
            # Обработка годов для уникальных цитирований
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, 2026)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                year_counts_unique = years_unique.value_counts().reset_index()
                year_counts_unique.columns = ['year', 'frequency_unique']
                year_counts_unique['percentage_unique'] = round(year_counts_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                year_counts_unique = pd.DataFrame(columns=['year', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем результаты
            if not year_counts_total.empty and not year_counts_unique.empty:
                year_counts = year_counts_total.merge(year_counts_unique, on='year', how='outer').fillna(0)
            elif not year_counts_total.empty:
                year_counts = year_counts_total.copy()
                year_counts['frequency_unique'] = 0
                year_counts['percentage_unique'] = 0.0
            elif not year_counts_unique.empty:
                year_counts = year_counts_unique.copy()
                year_counts['frequency_total'] = 0
                year_counts['percentage_total'] = 0.0
            else:
                return pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])
    
            return year_counts[['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('year', ascending=False)
        except Exception as e:
            st.error(f"Error in analyze_citation_year_distribution: {e}")
            return pd.DataFrame()
    
    def analyze_citation_five_year_periods(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of five-year periods for citing articles (from new to old)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()
    
            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)
    
            start_year = 1900
            current_year = datetime.now().year + 4
            period_starts = list(range(start_year, current_year + 1, 5))
            bins = period_starts + [period_starts[-1] + 5]
            labels = [f"{s}-{s+4}" for s in period_starts]
    
            # Обработка периодов для всех цитирований
            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                period_counts_total = pd.cut(years_total, bins=bins, labels=labels, right=False)
                period_df_total = period_counts_total.value_counts().reset_index()
                period_df_total.columns = ['period', 'frequency_total']
                period_df_total['percentage_total'] = round(period_df_total['frequency_total'] / total_citations * 100, 2)
            else:
                period_df_total = pd.DataFrame(columns=['period', 'frequency_total', 'percentage_total'])
    
            # Обработка периодов для уникальных цитирований
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, current_year)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                period_counts_unique = pd.cut(years_unique, bins=bins, labels=labels, right=False)
                period_df_unique = period_counts_unique.value_counts().reset_index()
                period_df_unique.columns = ['period', 'frequency_unique']
                period_df_unique['percentage_unique'] = round(period_df_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                period_df_unique = pd.DataFrame(columns=['period', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем результаты
            if not period_df_total.empty and not period_df_unique.empty:
                period_df = period_df_total.merge(period_df_unique, on='period', how='outer').fillna(0)
            elif not period_df_total.empty:
                period_df = period_df_total.copy()
                period_df['frequency_unique'] = 0
                period_df['percentage_unique'] = 0.0
            elif not period_df_unique.empty:
                period_df = period_df_unique.copy()
                period_df['frequency_total'] = 0
                period_df['percentage_total'] = 0.0
            else:
                return pd.DataFrame(columns=['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])
    
            return period_df[['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('period', ascending=False)
        except Exception as e:
            st.error(f"Error in analyze_citation_five_year_periods: {e}")
            return pd.DataFrame()

    def save_citation_analysis_to_excel(self, citing_articles_df: pd.DataFrame, citing_details_df: pd.DataFrame,
                                      doi_list: List[str], citing_results: Dict, all_citing_titles: List[str]) -> str:
        """Saves complete citing articles analysis to Excel"""
        try:
            timestamp = int(time.time())
            temp_dir = tempfile.mkdtemp()

            excel_path = os.path.join(tempfile.gettempdir(), f"citation_analysis_results_{timestamp}.xlsx")
            wb = Workbook()

            # First create Report_Summary tab
            ws_summary = wb.active
            ws_summary.title = 'Report_Summary_Citations'

            wb.remove(wb.active)

            try:
                unique_citations_df = self.get_unique_citations(citing_articles_df)
            except Exception as e:
                unique_citations_df = pd.DataFrame()

            try:
                duplicate_citations_df = self.find_duplicate_citations(citing_articles_df)
            except Exception as e:
                duplicate_citations_df = pd.DataFrame()

            stats = self.performance_monitor.get_stats()

            try:
                content_freq, compound_freq, scientific_freq = self.analyze_titles(all_citing_titles)
            except Exception as e:
                content_freq, compound_freq, scientific_freq = Counter(), Counter(), Counter()

            try:
                total_citation_relationships = len(citing_articles_df) if not citing_articles_df.empty else 0
                total_unique_citations = len(unique_citations_df) if not unique_citations_df.empty else 0

                countries_with_data = citing_articles_df[citing_articles_df['countries'].isin(['Unknown', 'Error']) == False]
                countries_percentage = (len(countries_with_data) / total_citation_relationships * 100) if total_citation_relationships > 0 else 0

                affiliations_with_data = citing_articles_df[citing_articles_df['affiliations'].isin(['Unknown', 'Error']) == False]
                affiliations_percentage = (len(affiliations_with_data) / total_citation_relationships * 100) if total_citation_relationships > 0 else 0

            except Exception as e:
                total_citation_relationships = 0
                total_unique_citations = 0
                countries_percentage = 0
                affiliations_percentage = 0

            citing_info = ""
            if citing_results:
                citing_info = f"\nCitations per source article:"
                for doi, data in citing_results.items():
                    citing_info += f"\n  - {doi}: {data['count']} citations"

            summary_content = f"""@MedvDmitry production

CITATION ANALYSIS REPORT (CITING ARTICLES)

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

ANALYSIS OVERVIEW
=================
Total source articles: {len(doi_list)}
Total citation relationships: {total_citation_relationships}
Total unique citing articles: {total_unique_citations}
Successful citations: {len(citing_articles_df[citing_articles_df['error'].isna()]) if not citing_articles_df.empty else 0}
Failed citations: {len(citing_articles_df[citing_articles_df['error'].notna()]) if not citing_articles_df.empty else 0}
Unique authors: {len(self.analyze_citation_authors_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique journals: {len(self.analyze_citation_journals_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique affiliations: {len(self.analyze_citation_affiliations_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Unique countries: {len(self.analyze_citation_countries_frequency(citing_articles_df)) if not citing_articles_df.empty else 0}
Duplicate citations: {len(duplicate_citations_df) if not duplicate_citations_df.empty else 0}

DATA COMPLETENESS
=================
Articles with country data: {countries_percentage:.1f}%
Articles with affiliation data: {affiliations_percentage:.1f}%

AFFILIATION PROCESSING
======================
Affiliations normalized and grouped by organization
Similar affiliations merged together
Frequency counts reflect grouped organizations

PERFORMANCE STATISTICS
======================
Total processing time: {stats.get('elapsed_seconds', 0):.2f} seconds ({stats.get('elapsed_minutes', 0):.2f} minutes)
Total API requests: {stats.get('total_requests', 0)}
Requests per second: {stats.get('requests_per_second', 0):.2f}
{citing_info}

DATA QUALITY NOTES
==================
Analysis focuses on articles that cite the source articles
Combined data from Crossref and OpenAlex improves completeness
All standard statistical analyses performed (authors, journals, countries, etc.)
Error handling ensures report generation even with partial data
Duplicate citations show articles that cite multiple source articles
Affiliations normalized and grouped for consistent organization names
"""

            # Create Report_Summary tab first
            ws_summary = wb.create_sheet('Report_Summary_Citations')
            for line in summary_content.split('\n'):
                ws_summary.append([line])

            # Remove authors_surnames column and add author_count for main tables
            if not citing_articles_df.empty and 'authors_surnames' in citing_articles_df.columns:
                citing_articles_df = citing_articles_df.drop(columns=['authors_surnames'])
            if not unique_citations_df.empty and 'authors_surnames' in unique_citations_df.columns:
                unique_citations_df = unique_citations_df.drop(columns=['authors_surnames'])
            if not duplicate_citations_df.empty and 'authors_surnames' in duplicate_citations_df.columns:
                duplicate_citations_df = duplicate_citations_df.drop(columns=['authors_surnames'])

            sheets_data = [
                ('Source_Articles_Citations', citing_details_df),
                ('All_Citations', citing_articles_df),
                ('All_Unique_Citations', unique_citations_df),
                ('Duplicate_Citations', duplicate_citations_df)
            ]

            analysis_methods = [
                ('Author_Frequency_Citations', self.analyze_citation_authors_frequency),
                ('Journal_Frequency_Citations', self.analyze_citation_journals_frequency),
                ('Affiliation_Frequency_Citations', self.analyze_citation_affiliations_frequency),
                ('Country_Frequency_Citations', self.analyze_citation_countries_frequency),
                ('Year_Distribution_Citations', self.analyze_citation_year_distribution),
                ('5_Years_Period_Citations', self.analyze_citation_five_year_periods)
            ]

            for sheet_name, method in analysis_methods:
                try:
                    result_df = method(citing_articles_df)
                    sheets_data.append((sheet_name, result_df))
                except Exception as e:
                    sheets_data.append((sheet_name, pd.DataFrame()))

            try:
                title_word_data = []
                for i, (word, count) in enumerate(content_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Content_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(compound_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Compound_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(scientific_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Scientific_Stopwords', 'Rank': i, 'Word': word, 'Frequency': count})

                title_word_df = pd.DataFrame(title_word_data)
                sheets_data.append(('Title_Word_Frequency_Citations', title_word_df))
            except Exception as e:
                sheets_data.append(('Title_Word_Frequency_Citations', pd.DataFrame()))

            for sheet_name, df in sheets_data:
                try:
                    if not df.empty:
                        ws = wb.create_sheet(sheet_name)
                        for r in dataframe_to_rows(df, index=False, header=True):
                            ws.append(r)
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([f"No data available for {sheet_name}"])
                except Exception as e:
                    pass

            wb.save(excel_path)

            return excel_path

        except Exception as e:
            try:
                timestamp = int(time.time())
                excel_path = os.path.join(tempfile.gettempdir(), f"minimal_citation_analysis_results_{timestamp}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Error_Report_Citations"
                ws.append(["ERROR REPORT - CITATION ANALYSIS"])
                ws.append([f"Critical error during citation analysis: {str(e)}"])
                ws.append([f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws.append(["DOIs processed:", ', '.join(doi_list)])
                wb.save(excel_path)
                return excel_path
            except:
                return "error_creating_citation_report"

    def process_doi_sequential(self, doi_list: List[str]) -> tuple[pd.DataFrame, pd.DataFrame, int, int, List[str]]:
        """Process DOIs sequentially to avoid API overload"""
        self.performance_monitor.start()

        all_references = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, doi in enumerate(doi_list):
            try:
                status_text.text(f"Collecting references for DOI {i+1}/{len(doi_list)}: {doi}")
                references = self.get_references_from_crossref(doi)
                for j, ref in enumerate(references):
                    all_references.append({
                        'source_doi': doi,
                        'position': j + 1,
                        'ref': ref
                    })
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
                
                # Update progress
                progress_bar.progress((i + 1) / len(doi_list))
            except Exception as e:
                pass

        status_text.empty()
        progress_bar.empty()

        unique_dois = set()
        titles_to_search = set()
        all_titles = []

        for ref_data in all_references:
            ref = ref_data['ref']
            ref_doi = ref.get('DOI')
            title = ref.get('article-title', 'Unknown')
            all_titles.append(title)

            if ref_doi and self.validate_doi(ref_doi):
                unique_dois.add(ref_doi)
            elif title != 'Unknown':
                titles_to_search.add(title)

        title_to_doi = {}
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        titles_list = list(titles_to_search)
        for i, title in enumerate(titles_list):
            status_text.text(f"Searching DOIs by title {i+1}/{len(titles_list)}")
            doi = self.quick_doi_search(title)
            if doi and self.validate_doi(doi):
                normalized_doi = self.normalize_doi(doi)
                title_to_doi[title] = normalized_doi
                unique_dois.add(normalized_doi)
            time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            
            # Update progress
            progress_bar.progress((i + 1) / len(titles_list))

        status_text.empty()
        progress_bar.empty()

        progress_bar = st.progress(0)
        status_text = st.empty()
        
        unique_dois_list = list(unique_dois)
        for i, doi in enumerate(unique_dois_list):
            status_text.text(f"Processing unique DOI {i+1}/{len(unique_dois_list)}: {doi}")
            if doi not in self.unique_ref_data_cache:
                try:
                    article_data = self.get_combined_article_data(doi)
                    self.unique_ref_data_cache[doi] = {
                        'doi': doi,
                        'title': article_data['title'],
                        'authors': article_data['authors'],
                        'authors_surnames': article_data['authors_surnames'],
                        'authors_with_initials': article_data['authors_with_initials'],
                        'author_count': article_data['author_count'],
                        'year': article_data['year'],
                        'journal_full_name': article_data['journal_full_name'],
                        'journal_abbreviation': article_data['journal_abbreviation'],
                        'publisher': article_data['publisher'],
                        'citation_count_crossref': article_data['citation_count_crossref'],
                        'citation_count_openalex': article_data['citation_count_openalex'],
                        'affiliations': article_data['affiliations'],
                        'countries': article_data['countries'],
                        'publication_year': article_data.get('publication_year'),
                        'years_since_publication': article_data['years_since_publication']
                    }
                except Exception as e:
                    self.unique_ref_data_cache[doi] = {
                        'doi': doi, 'title': 'Unknown', 'authors': 'Error',
                        'authors_surnames': 'Error', 'authors_with_initials': 'Error', 'author_count': 0,
                        'year': 'Unknown', 'journal_full_name': 'Error',
                        'journal_abbreviation': 'Error', 'publisher': 'Error',
                        'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                        'affiliations': 'Error', 'countries': 'Error', 'error': str(e)
                    }
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
                
            # Update progress
            progress_bar.progress((i + 1) / len(unique_dois_list))

        status_text.empty()
        progress_bar.empty()

        results = []
        source_articles = []

        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, doi in enumerate(doi_list):
            try:
                status_text.text(f"Processing source article {i+1}/{len(doi_list)}: {doi}")
                source_data = self.get_combined_article_data(doi)
                source_row = {
                    'source_doi': doi,
                    'position': None,
                    'doi': doi,
                    'title': source_data['title'],
                    'authors': source_data['authors'],
                    'authors_with_initials': source_data['authors_with_initials'],
                    'author_count': source_data['author_count'],
                    'year': source_data['year'],
                    'journal_full_name': source_data['journal_full_name'],
                    'journal_abbreviation': source_data['journal_abbreviation'],
                    'publisher': source_data['publisher'],
                    'citation_count_crossref': source_data['citation_count_crossref'],
                    'citation_count_openalex': source_data['citation_count_openalex'],
                    'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                        source_data['citation_count_crossref'], source_data.get('publication_year')
                    ),
                    'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                        source_data['citation_count_openalex'], source_data.get('publication_year')
                    ),
                    'years_since_publication': source_data['years_since_publication'],
                    'affiliations': source_data['affiliations'],
                    'countries': source_data['countries'],
                    'error': None
                }
                source_articles.append(source_row)
            except Exception as e:
                source_articles.append({
                    'source_doi': doi, 'position': None, 'doi': doi, 'title': 'Unknown',
                    'authors': 'Error', 'authors_with_initials': 'Error', 'author_count': 0,
                    'year': 'Unknown', 'journal_full_name': 'Error', 'journal_abbreviation': 'Error',
                    'publisher': 'Error', 'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                    'annual_citation_rate_crossref': 'N/A', 'annual_citation_rate_openalex': 'N/A',
                    'years_since_publication': 'N/A', 'affiliations': 'Error', 'countries': 'Error',
                    'error': str(e)
                })

            article_refs = [ref for ref in all_references if ref['source_doi'] == doi]
            for ref_data in article_refs:
                ref = ref_data['ref']
                position = ref_data['position']
                ref_doi = ref.get('DOI')
                title = ref.get('article-title', 'Unknown')

                if ref_doi and self.validate_doi(ref_doi) and ref_doi in self.unique_ref_data_cache:
                    ref_info = self.unique_ref_data_cache[ref_doi].copy()
                    ref_row = {
                        'source_doi': doi,
                        'position': position,
                        'doi': ref_doi,
                        'title': ref_info['title'],
                        'authors': ref_info['authors'],
                        'authors_with_initials': ref_info['authors_with_initials'],
                        'author_count': ref_info['author_count'],
                        'year': ref_info['year'],
                        'journal_full_name': ref_info['journal_full_name'],
                        'journal_abbreviation': ref_info['journal_abbreviation'],
                        'publisher': ref_info['publisher'],
                        'citation_count_crossref': ref_info['citation_count_crossref'],
                        'citation_count_openalex': ref_info['citation_count_openalex'],
                        'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                            ref_info['citation_count_crossref'], ref_info.get('publication_year')
                        ),
                        'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                            ref_info['citation_count_openalex'], ref_info.get('publication_year')
                        ),
                        'years_since_publication': ref_info['years_since_publication'],
                        'affiliations': ref_info['affiliations'],
                        'countries': ref_info['countries'],
                        'error': None
                    }
                    results.append(ref_row)
                else:
                    found_doi = title_to_doi.get(title)
                    if found_doi and found_doi in self.unique_ref_data_cache:
                        ref_info = self.unique_ref_data_cache[found_doi].copy()
                        ref_row = {
                            'source_doi': doi,
                            'position': position,
                            'doi': found_doi,
                            'title': ref_info['title'],
                            'authors': ref_info['authors'],
                            'authors_with_initials': ref_info['authors_with_initials'],
                            'author_count': ref_info['author_count'],
                            'year': ref_info['year'],
                            'journal_full_name': ref_info['journal_full_name'],
                            'journal_abbreviation': ref_info['journal_abbreviation'],
                            'publisher': ref_info['publisher'],
                            'citation_count_crossref': ref_info['citation_count_crossref'],
                            'citation_count_openalex': ref_info['citation_count_openalex'],
                            'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                                ref_info['citation_count_crossref'], ref_info.get('publication_year')
                            ),
                            'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                                ref_info['citation_count_openalex'], ref_info.get('publication_year')
                            ),
                            'years_since_publication': ref_info['years_since_publication'],
                            'affiliations': ref_info['affiliations'],
                            'countries': ref_info['countries'],
                            'error': None
                        }
                        results.append(ref_row)
                    else:
                        results.append({
                            'source_doi': doi, 'position': position, 'doi': ref_doi, 'title': title,
                            'authors': 'Unknown', 'authors_with_initials': 'Unknown', 'author_count': 0,
                            'year': ref.get('year', 'Unknown'), 'journal_full_name': 'Unknown',
                            'journal_abbreviation': 'Unknown', 'publisher': 'Unknown',
                            'citation_count_crossref': 'N/A', 'citation_count_openalex': 'N/A',
                            'annual_citation_rate_crossref': 'N/A', 'annual_citation_rate_openalex': 'N/A',
                            'years_since_publication': 'N/A', 'affiliations': 'Unknown', 'countries': 'Unknown',
                            'error': f"Invalid or missing DOI: {ref_doi}, no match found for title '{title}'"
                        })

            time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            
            # Update progress
            progress_bar.progress((i + 1) / len(doi_list))

        status_text.empty()
        progress_bar.empty()

        try:
            combined_references_df = pd.DataFrame(results)
        except Exception as e:
            combined_references_df = pd.DataFrame()

        try:
            source_articles_df = pd.DataFrame(source_articles)
        except Exception as e:
            source_articles_df = pd.DataFrame()

        try:
            combined_references_df = self.enhance_incomplete_data(combined_references_df)
        except Exception as e:
            pass

        try:
            source_articles_df = self.enhance_incomplete_data(source_articles_df)
        except Exception as e:
            pass

        return combined_references_df, source_articles_df, len(all_references), len(unique_dois), all_titles

    def enhance_incomplete_data(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Enhance references with incomplete data"""
        if references_df.empty:
            return references_df

        enhanced_rows = []
        incomplete_count = 0

        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for index, row in enumerate(references_df.iterrows()):
            i, row_data = row
            status_text.text(f"Enhancing data {i+1}/{len(references_df)}")
            doi = row_data['doi']

            needs_enhancement = (
                pd.isna(doi) or
                row_data['title'] == 'Unknown' or
                row_data['authors'] == 'Unknown' or
                row_data['affiliations'] == 'Unknown' or
                row_data['countries'] == 'Unknown' or
                pd.notna(row_data.get('error'))
            )

            if needs_enhancement and doi and self.validate_doi(doi):
                incomplete_count += 1
                try:
                    enhanced_data = self.get_combined_article_data(doi)
                    enhanced_row = {
                        'source_doi': row_data['source_doi'],
                        'position': row_data['position'],
                        'doi': doi,
                        'title': enhanced_data['title'],
                        'authors': enhanced_data['authors'],
                        'authors_with_initials': enhanced_data['authors_with_initials'],
                        'author_count': enhanced_data['author_count'],
                        'year': enhanced_data['year'],
                        'journal_full_name': enhanced_data['journal_full_name'],
                        'journal_abbreviation': enhanced_data['journal_abbreviation'],
                        'publisher': enhanced_data['publisher'],
                        'citation_count_crossref': enhanced_data['citation_count_crossref'],
                        'citation_count_openalex': enhanced_data['citation_count_openalex'],
                        'annual_citation_rate_crossref': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_crossref'], enhanced_data.get('publication_year')
                        ),
                        'annual_citation_rate_openalex': self.safe_calculate_annual_citation_rate(
                            enhanced_data['citation_count_openalex'], enhanced_data.get('publication_year')
                        ),
                        'years_since_publication': enhanced_data['years_since_publication'],
                        'affiliations': enhanced_data['affiliations'],
                        'countries': enhanced_data['countries'],
                        'error': None
                    }
                    enhanced_rows.append(enhanced_row)
                    time.sleep(Config.DELAY_BETWEEN_REQUESTS)
                    continue
                except Exception as e:
                    pass

            enhanced_rows.append(row_data.to_dict())
            
            # Update progress
            progress_bar.progress((i + 1) / len(references_df))

        status_text.empty()
        progress_bar.empty()

        return pd.DataFrame(enhanced_rows)

    def get_unique_references(self, references_df: pd.DataFrame) -> pd.DataFrame:
        if references_df.empty:
            return pd.DataFrame()

        cache_key = id(references_df)
        if cache_key not in self._unique_references_cache:
            references_df['ref_id'] = references_df['doi'].fillna('') + '|' + references_df['title'].fillna('')
            unique_df = references_df.drop_duplicates(subset=['ref_id'], keep='first').drop(columns=['ref_id'])
            self._unique_references_cache[cache_key] = unique_df
        return self._unique_references_cache[cache_key]

    def analyze_authors_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            authors_total = references_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_total = authors_total[authors_total.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_total = authors_total.value_counts().reset_index()
            author_freq_total.columns = ['author_with_initial', 'frequency_total']
            author_freq_total['percentage_total'] = round(author_freq_total['frequency_total'] / total_refs * 100, 2)
            authors_unique = unique_df['authors_with_initials'].str.split(',', expand=True).stack()
            authors_unique = authors_unique[authors_unique.str.strip().isin(['Unknown', 'Error']) == False]
            author_freq_unique = authors_unique.value_counts().reset_index()
            author_freq_unique.columns = ['author_with_initial', 'frequency_unique']
            author_freq_unique['percentage_unique'] = round(author_freq_unique['frequency_unique'] / total_unique * 100, 2)
            author_freq = author_freq_total.merge(author_freq_unique, on='author_with_initial', how='outer').fillna(0)
            return author_freq[['author_with_initial', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_journals_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            journals_total = references_df['journal_abbreviation']
            journals_total = journals_total[journals_total.isin(['Unknown', 'Error']) == False]
            journal_freq_total = journals_total.value_counts().reset_index()
            journal_freq_total.columns = ['journal_abbreviation', 'frequency_total']
            journal_freq_total['percentage_total'] = round(journal_freq_total['frequency_total'] / total_refs * 100, 2)
            journals_unique = unique_df['journal_abbreviation']
            journals_unique = journals_unique[journals_unique.isin(['Unknown', 'Error']) == False]
            journal_freq_unique = journals_unique.value_counts().reset_index()
            journal_freq_unique.columns = ['journal_abbreviation', 'frequency_unique']
            journal_freq_unique['percentage_unique'] = round(journal_freq_unique['frequency_unique'] / total_unique * 100, 2)
            journal_freq = journal_freq_total.merge(journal_freq_unique, on='journal_abbreviation', how='outer').fillna(0)

            # Add additional metrics for citations
            journal_citation_metrics = []
            for journal in journal_freq['journal_abbreviation']:
                journal_articles = unique_df[unique_df['journal_abbreviation'] == journal]

                total_crossref_citations = journal_articles['citation_count_crossref'].sum()
                total_openalex_citations = journal_articles['citation_count_openalex'].sum()

                avg_crossref_citations = journal_articles['citation_count_crossref'].mean() if len(journal_articles) > 0 else 0
                avg_openalex_citations = journal_articles['citation_count_openalex'].mean() if len(journal_articles) > 0 else 0

                journal_citation_metrics.append({
                    'journal_abbreviation': journal,
                    'total_crossref_citations': total_crossref_citations,
                    'total_openalex_citations': total_openalex_citations,
                    'avg_crossref_citations': round(avg_crossref_citations, 2),
                    'avg_openalex_citations': round(avg_openalex_citations, 2)
                })

            journal_metrics_df = pd.DataFrame(journal_citation_metrics)
            journal_freq = journal_freq.merge(journal_metrics_df, on='journal_abbreviation', how='left').fillna(0)

            return journal_freq[['journal_abbreviation', 'frequency_total', 'percentage_total',
                               'frequency_unique', 'percentage_unique', 'total_crossref_citations',
                               'total_openalex_citations', 'avg_crossref_citations', 'avg_openalex_citations']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_affiliations_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of affiliation frequency in references with grouping"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)

            all_affiliations = []
            for affil_string in references_df['affiliations']:
                if pd.isna(affil_string) or affil_string in ['Unknown', 'Error', '']:
                    continue

                try:
                    affil_list = affil_string.split(';')
                    for affil in affil_list:
                        clean_affil = affil.strip()
                        if clean_affil and clean_affil not in ['Unknown', 'Error']:
                            all_affiliations.append(clean_affil)
                except Exception:
                    pass

            if not all_affiliations:
                return pd.DataFrame(columns=['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])

            affiliation_frequencies, grouped_organizations = self.fast_affiliation_processor.process_affiliations_list_fast(all_affiliations)

            affil_data = []
            for affil, freq in affiliation_frequencies.items():
                percentage_total = round(freq / total_refs * 100, 2) if total_refs > 0 else 0
                affil_data.append({
                    'affiliation': affil,
                    'frequency_total': freq,
                    'percentage_total': percentage_total
                })

            unique_affiliations = []
            for affil_string in unique_df['affiliations']:
                if affil_string and affil_string not in ['Unknown', 'Error']:
                    affil_list = affil_string.split(';')
                    unique_affiliations.extend([affil.strip() for affil in affil_list if affil.strip()])

            if unique_affiliations:
                unique_frequencies, _ = self.fast_affiliation_processor.process_affiliations_list_fast(unique_affiliations)

                affil_df = pd.DataFrame(affil_data)
                for affil, freq in unique_frequencies.items():
                    percentage_unique = round(freq / total_unique * 100, 2) if total_unique > 0 else 0
                    if affil in affil_df['affiliation'].values:
                        affil_df.loc[affil_df['affiliation'] == affil, 'frequency_unique'] = freq
                        affil_df.loc[affil_df['affiliation'] == affil, 'percentage_unique'] = percentage_unique
                    else:
                        affil_df = pd.concat([affil_df, pd.DataFrame([{
                            'affiliation': affil,
                            'frequency_total': 0,
                            'percentage_total': 0,
                            'frequency_unique': freq,
                            'percentage_unique': percentage_unique
                        }])], ignore_index=True)
            else:
                affil_df = pd.DataFrame(affil_data)
                affil_df['frequency_unique'] = affil_df['frequency_total']
                affil_df['percentage_unique'] = affil_df['percentage_total']

            affil_df = affil_df.fillna(0)
            return affil_df[['affiliation', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)

        except Exception as e:
            return pd.DataFrame()

    def analyze_countries_frequency(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of country frequency in references with separation into individual countries and collaborations"""
        try:
            if references_df.empty:
                return pd.DataFrame()

            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)

            # Analysis of individual countries
            single_countries = []
            collaborations = []

            for countries in references_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries.extend(country_list)
                    else:
                        collaborations.append(countries)

            # Frequency of individual countries
            single_country_counter = Counter(single_countries)
            single_country_freq = pd.DataFrame({
                'countries': list(single_country_counter.keys()),
                'type': ['single'] * len(single_country_counter),
                'frequency_total': list(single_country_counter.values())
            })
            single_country_freq['percentage_total'] = round(single_country_freq['frequency_total'] / total_refs * 100, 2)

            # Frequency of collaborations
            collaboration_counter = Counter(collaborations)
            collaboration_freq = pd.DataFrame({
                'countries': list(collaboration_counter.keys()),
                'type': ['collaboration'] * len(collaboration_counter),
                'frequency_total': list(collaboration_counter.values())
            })
            collaboration_freq['percentage_total'] = round(collaboration_freq['frequency_total'] / total_refs * 100, 2)

            # Combine
            country_freq_total = pd.concat([single_country_freq, collaboration_freq], ignore_index=True)

            # Similarly for unique articles
            single_countries_unique = []
            collaborations_unique = []

            for countries in unique_df['countries']:
                if countries not in ['Unknown', 'Error']:
                    country_list = [c.strip() for c in countries.split(';')]
                    if len(country_list) == 1:
                        single_countries_unique.extend(country_list)
                    else:
                        collaborations_unique.append(countries)

            single_country_counter_unique = Counter(single_countries_unique)
            collaboration_counter_unique = Counter(collaborations_unique)

            single_country_freq_unique = pd.DataFrame({
                'countries': list(single_country_counter_unique.keys()),
                'frequency_unique': list(single_country_counter_unique.values())
            })
            single_country_freq_unique['percentage_unique'] = round(single_country_freq_unique['frequency_unique'] / total_unique * 100, 2)

            collaboration_freq_unique = pd.DataFrame({
                'countries': list(collaboration_counter_unique.keys()),
                'frequency_unique': list(collaboration_counter_unique.values())
            })
            collaboration_freq_unique['percentage_unique'] = round(collaboration_freq_unique['frequency_unique'] / total_unique * 100, 2)

            country_freq_unique = pd.concat([
                single_country_freq_unique.assign(type='single'),
                collaboration_freq_unique.assign(type='collaboration')
            ], ignore_index=True)

            country_freq = country_freq_total.merge(country_freq_unique, on=['countries', 'type'], how='outer').fillna(0)

            # Add full country names
            country_freq['country_fullname'] = country_freq['countries'].apply(
                lambda x: ';'.join([self.fast_affiliation_processor.get_country_display(code) for code in x.split(';')])
            )

            return country_freq[['countries', 'country_fullname', 'type', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']].sort_values('frequency_total', ascending=False)
        except Exception as e:
            return pd.DataFrame()

    def analyze_year_distribution(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of year distribution for references (from new to old)"""
        try:
            if references_df.empty:
                return pd.DataFrame()
    
            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
            
            # Получаем все уникальные годы из данных
            all_years_total = pd.to_numeric(references_df['year'], errors='coerce')
            all_years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            
            # Объединяем все годы и фильтруем корректные
            all_years = pd.concat([all_years_total, all_years_unique])
            valid_years = all_years[all_years.notna() & all_years.between(1900, 2026)].astype(int).unique()
            
            if len(valid_years) == 0:
                return pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])
            
            # Создаем базовый DataFrame со всеми годами
            all_years_df = pd.DataFrame({'year': sorted(valid_years, reverse=True)})
    
            # Обработка годов для всех ссылок
            years_total = pd.to_numeric(references_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, 2026)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                year_counts_total = years_total.value_counts().reset_index()
                year_counts_total.columns = ['year', 'frequency_total']
                year_counts_total['percentage_total'] = round(year_counts_total['frequency_total'] / total_refs * 100, 2)
            else:
                year_counts_total = pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total'])
    
            # Обработка годов для уникальных ссылок
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, 2026)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                year_counts_unique = years_unique.value_counts().reset_index()
                year_counts_unique.columns = ['year', 'frequency_unique']
                year_counts_unique['percentage_unique'] = round(year_counts_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                year_counts_unique = pd.DataFrame(columns=['year', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем с базовым DataFrame
            year_df = all_years_df.copy()
            
            if not year_counts_total.empty:
                year_df = year_df.merge(year_counts_total, on='year', how='left')
            else:
                year_df['frequency_total'] = 0
                year_df['percentage_total'] = 0.0
                
            if not year_counts_unique.empty:
                year_df = year_df.merge(year_counts_unique, on='year', how='left')
            else:
                year_df['frequency_unique'] = 0
                year_df['percentage_unique'] = 0.0
    
            # Заполняем NaN значения нулями
            year_df = year_df.fillna(0)
            
            return year_df[['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']]
        except Exception as e:
            st.error(f"Error in analyze_year_distribution: {e}")
            return pd.DataFrame()
    
    def analyze_citation_year_distribution(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of year distribution for citing articles (from new to old)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()
    
            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)
            
            # Получаем все уникальные годы из данных
            all_years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            all_years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            
            # Объединяем все годы и фильтруем корректные
            all_years = pd.concat([all_years_total, all_years_unique])
            valid_years = all_years[all_years.notna() & all_years.between(1900, 2026)].astype(int).unique()
            
            if len(valid_years) == 0:
                return pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique'])
            
            # Создаем базовый DataFrame со всеми годами
            all_years_df = pd.DataFrame({'year': sorted(valid_years, reverse=True)})
    
            # Обработка годов для всех цитирований
            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, 2026)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                year_counts_total = years_total.value_counts().reset_index()
                year_counts_total.columns = ['year', 'frequency_total']
                year_counts_total['percentage_total'] = round(year_counts_total['frequency_total'] / total_citations * 100, 2)
            else:
                year_counts_total = pd.DataFrame(columns=['year', 'frequency_total', 'percentage_total'])
    
            # Обработка годов для уникальных цитирований
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, 2026)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                year_counts_unique = years_unique.value_counts().reset_index()
                year_counts_unique.columns = ['year', 'frequency_unique']
                year_counts_unique['percentage_unique'] = round(year_counts_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                year_counts_unique = pd.DataFrame(columns=['year', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем с базовым DataFrame
            year_df = all_years_df.copy()
            
            if not year_counts_total.empty:
                year_df = year_df.merge(year_counts_total, on='year', how='left')
            else:
                year_df['frequency_total'] = 0
                year_df['percentage_total'] = 0.0
                
            if not year_counts_unique.empty:
                year_df = year_df.merge(year_counts_unique, on='year', how='left')
            else:
                year_df['frequency_unique'] = 0
                year_df['percentage_unique'] = 0.0
    
            # Заполняем NaN значения нулями
            year_df = year_df.fillna(0)
            
            return year_df[['year', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']]
        except Exception as e:
            st.error(f"Error in analyze_citation_year_distribution: {e}")
            return pd.DataFrame()
    
    def analyze_five_year_periods(self, references_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of five-year periods for references (from new to old)"""
        try:
            if references_df.empty:
                return pd.DataFrame()
    
            total_refs = len(references_df)
            unique_df = self.get_unique_references(references_df)
            total_unique = len(unique_df)
    
            start_year = 1900
            current_year = datetime.now().year + 4
            period_starts = list(range(start_year, current_year + 1, 5))
            bins = period_starts + [period_starts[-1] + 5]
            labels = [f"{s}-{s+4}" for s in period_starts]
    
            # Создаем базовый DataFrame со всеми периодами
            all_periods_df = pd.DataFrame({'period': labels})
    
            # Обработка периодов для всех ссылок
            years_total = pd.to_numeric(references_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                period_counts_total = pd.cut(years_total, bins=bins, labels=labels, right=False)
                period_df_total = period_counts_total.value_counts().reset_index()
                period_df_total.columns = ['period', 'frequency_total']
                period_df_total['percentage_total'] = round(period_df_total['frequency_total'] / total_refs * 100, 2)
            else:
                period_df_total = pd.DataFrame(columns=['period', 'frequency_total', 'percentage_total'])
    
            # Обработка периодов для уникальных ссылок
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, current_year)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                period_counts_unique = pd.cut(years_unique, bins=bins, labels=labels, right=False)
                period_df_unique = period_counts_unique.value_counts().reset_index()
                period_df_unique.columns = ['period', 'frequency_unique']
                period_df_unique['percentage_unique'] = round(period_df_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                period_df_unique = pd.DataFrame(columns=['period', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем с базовым DataFrame, чтобы включить все периоды
            period_df = all_periods_df.copy()
            
            if not period_df_total.empty:
                period_df = period_df.merge(period_df_total, on='period', how='left')
            else:
                period_df['frequency_total'] = 0
                period_df['percentage_total'] = 0.0
                
            if not period_df_unique.empty:
                period_df = period_df.merge(period_df_unique, on='period', how='left')
            else:
                period_df['frequency_unique'] = 0
                period_df['percentage_unique'] = 0.0
    
            # Заполняем NaN значения нулями
            period_df = period_df.fillna(0)
            
            # Сортируем от новых к старым
            period_df = period_df.sort_values('period', ascending=False)
    
            return period_df[['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']]
        except Exception as e:
            st.error(f"Error in analyze_five_year_periods: {e}")
            return pd.DataFrame()
    
    def analyze_citation_five_year_periods(self, citations_df: pd.DataFrame) -> pd.DataFrame:
        """Analysis of five-year periods for citing articles (from new to old)"""
        try:
            if citations_df.empty:
                return pd.DataFrame()
    
            total_citations = len(citations_df)
            unique_df = self.get_unique_citations(citations_df)
            total_unique = len(unique_df)
    
            start_year = 1900
            current_year = datetime.now().year + 4
            period_starts = list(range(start_year, current_year + 1, 5))
            bins = period_starts + [period_starts[-1] + 5]
            labels = [f"{s}-{s+4}" for s in period_starts]
    
            # Создаем базовый DataFrame со всеми периодами
            all_periods_df = pd.DataFrame({'period': labels})
    
            # Обработка периодов для всех цитирований
            years_total = pd.to_numeric(citations_df['year'], errors='coerce')
            years_total = years_total[years_total.notna() & years_total.between(1900, current_year)]
            if not years_total.empty:
                years_total = years_total.astype(int)
                period_counts_total = pd.cut(years_total, bins=bins, labels=labels, right=False)
                period_df_total = period_counts_total.value_counts().reset_index()
                period_df_total.columns = ['period', 'frequency_total']
                period_df_total['percentage_total'] = round(period_df_total['frequency_total'] / total_citations * 100, 2)
            else:
                period_df_total = pd.DataFrame(columns=['period', 'frequency_total', 'percentage_total'])
    
            # Обработка периодов для уникальных цитирований
            years_unique = pd.to_numeric(unique_df['year'], errors='coerce')
            years_unique = years_unique[years_unique.notna() & years_unique.between(1900, current_year)]
            if not years_unique.empty:
                years_unique = years_unique.astype(int)
                period_counts_unique = pd.cut(years_unique, bins=bins, labels=labels, right=False)
                period_df_unique = period_counts_unique.value_counts().reset_index()
                period_df_unique.columns = ['period', 'frequency_unique']
                period_df_unique['percentage_unique'] = round(period_df_unique['frequency_unique'] / total_unique * 100, 2)
            else:
                period_df_unique = pd.DataFrame(columns=['period', 'frequency_unique', 'percentage_unique'])
    
            # Объединяем с базовым DataFrame, чтобы включить все периоды
            period_df = all_periods_df.copy()
            
            if not period_df_total.empty:
                period_df = period_df.merge(period_df_total, on='period', how='left')
            else:
                period_df['frequency_total'] = 0
                period_df['percentage_total'] = 0.0
                
            if not period_df_unique.empty:
                period_df = period_df.merge(period_df_unique, on='period', how='left')
            else:
                period_df['frequency_unique'] = 0
                period_df['percentage_unique'] = 0.0
    
            # Заполняем NaN значения нулями
            period_df = period_df.fillna(0)
            
            # Сортируем от новых к старым
            period_df = period_df.sort_values('period', ascending=False)
    
            return period_df[['period', 'frequency_total', 'percentage_total', 'frequency_unique', 'percentage_unique']]
        except Exception as e:
            st.error(f"Error in analyze_citation_five_year_periods: {e}")
            return pd.DataFrame()

    def find_duplicate_references(self, references_df: pd.DataFrame) -> pd.DataFrame:
        try:
            if references_df.empty:
                return pd.DataFrame()

            references_df['ref_id'] = references_df['doi'].fillna('') + '|' + references_df['title'].fillna('')
            ref_counts = references_df.groupby('ref_id')['source_doi'].nunique().reset_index()
            duplicate_ref_ids = ref_counts[ref_counts['source_doi'] > 1]['ref_id']
            if duplicate_ref_ids.empty:
                columns = list(references_df.columns) + ['frequency']
                columns.remove('ref_id')
                return pd.DataFrame(columns=columns)
            frequency_map = references_df['ref_id'].value_counts().to_dict()
            duplicates = references_df[references_df['ref_id'].isin(duplicate_ref_ids)].copy()
            duplicates = duplicates.drop_duplicates(subset=['ref_id'], keep='first')
            duplicates = duplicates[~((duplicates['doi'].isna()) & (duplicates['title'] == 'Unknown'))]
            duplicates['frequency'] = duplicates['ref_id'].map(frequency_map)
            duplicates = duplicates.drop(columns=['ref_id'])
            return duplicates.sort_values(['frequency', 'doi'], ascending=[False, True])
        except Exception as e:
            return pd.DataFrame()

    def preprocess_content_words(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        text = re.sub(r'[^a-zA-Z\s-]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        words = text.split()
        content_words = []
        for word in words:
            if '-' in word:
                continue
            if len(word) > 2 and word not in self.stop_words and word != 'sub':
                stemmed_word = self.stemmer.stem(word)
                if stemmed_word not in self.scientific_stopwords_stemmed:
                    content_words.append(stemmed_word)
        return content_words

    def extract_compound_words(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        compound_words = re.findall(r'\b[a-z]{2,}-[a-z]{2,}(?:-[a-z]{2,})*\b', text)
        return [word for word in compound_words if not any(part in self.stop_words for part in word.split('-'))]

    def extract_scientific_stopwords(self, text: str) -> List[str]:
        if not text or text in ['Unknown', 'Error']:
            return []
        text = text.lower()
        text = re.sub(r'[^a-zA-Z\s]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        words = text.split()
        scientific_words = []
        for word in words:
            if len(word) > 2:
                stemmed_word = self.stemmer.stem(word)
                if stemmed_word in self.scientific_stopwords_stemmed:
                    for original_word in self.scientific_stopwords:
                        if self.stemmer.stem(original_word) == stemmed_word:
                            scientific_words.append(original_word)
                            break
        return scientific_words

    def analyze_titles(self, titles: List[str]) -> tuple[Counter, Counter, Counter]:
        content_words = []
        compound_words = []
        scientific_words = []
        valid_titles = [t for t in titles if t not in ['Unknown', 'Error']]
        for title in valid_titles:
            content_words.extend(self.preprocess_content_words(title))
            compound_words.extend(self.extract_compound_words(title))
            scientific_words.extend(self.extract_scientific_stopwords(title))
        return Counter(content_words), Counter(compound_words), Counter(scientific_words)

    def save_all_data_to_excel(self, combined_df: pd.DataFrame, source_articles_df: pd.DataFrame,
                         doi_list: List[str], total_references: int, unique_dois: int,
                         all_titles: List[str]) -> str:
        """Saves references analysis to Excel"""
        try:
            timestamp = int(time.time())
            temp_dir = tempfile.mkdtemp()

            excel_path = os.path.join(tempfile.gettempdir(), f"references_analysis_results_{timestamp}.xlsx")
            wb = Workbook()

            # First create Report_Summary tab
            ws_summary = wb.active
            ws_summary.title = 'Report_Summary'

            wb.remove(wb.active)

            try:
                unique_df = self.get_unique_references(combined_df)
            except Exception as e:
                unique_df = pd.DataFrame()

            try:
                duplicate_df = self.find_duplicate_references(combined_df)
            except Exception as e:
                duplicate_df = pd.DataFrame()

            try:
                failed_df = combined_df[combined_df['error'].notna()][['source_doi', 'position', 'doi', 'error']].copy()
                failed_df.columns = ['source_doi', 'ref_number', 'reference_doi', 'error_description']
            except Exception as e:
                failed_df = pd.DataFrame()

            stats = self.performance_monitor.get_stats()

            try:
                content_freq, compound_freq, scientific_freq = self.analyze_titles(all_titles)
            except Exception as e:
                content_freq, compound_freq, scientific_freq = Counter(), Counter(), Counter()

            try:
                total_processed = len(combined_df) if not combined_df.empty else 0

                countries_with_data = combined_df[combined_df['countries'].isin(['Unknown', 'Error']) == False]
                countries_percentage = (len(countries_with_data) / total_processed * 100) if total_processed > 0 else 0

                affiliations_with_data = combined_df[combined_df['affiliations'].isin(['Unknown', 'Error']) == False]
                affiliations_percentage = (len(affiliations_with_data) / total_processed * 100) if total_processed > 0 else 0

            except Exception as e:
                countries_percentage = 0
                affiliations_percentage = 0

            summary_content = f"""@MedvDmitry production

REFERENCES ANALYSIS REPORT

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

ANALYSIS OVERVIEW
=================
Total source articles: {len(doi_list)}
Total references collected: {total_references}
Unique DOIs identified: {unique_dois}
Total references processed: {len(combined_df) if not combined_df.empty else 0}
Unique references: {len(unique_df) if not unique_df.empty else 0}
Successful references: {len(combined_df[combined_df['error'].isna()]) if not combined_df.empty else 0}
Failed references: {len(combined_df[combined_df['error'].notna()]) if not combined_df.empty else 0}
Unique authors: {len(self.analyze_authors_frequency(combined_df)) if not combined_df.empty else 0}
Unique journals: {len(self.analyze_journals_frequency(combined_df)) if not combined_df.empty else 0}
Unique affiliations: {len(self.analyze_affiliations_frequency(combined_df)) if not combined_df.empty else 0}
Unique countries: {len(self.analyze_countries_frequency(combined_df)) if not combined_df.empty else 0}
Duplicate references: {len(duplicate_df) if not duplicate_df.empty else 0}

DATA COMPLETENESS
=================
References with country data: {countries_percentage:.1f}%
References with affiliation data: {affiliations_percentage:.1f}%

AFFILIATION PROCESSING
======================
Affiliations normalized and grouped by organization
Similar affiliations merged together
Frequency counts reflect grouped organizations

PERFORMANCE STATISTICS
======================
Total processing time: {stats.get('elapsed_seconds', 0):.2f} seconds ({stats.get('elapsed_minutes', 0):.2f} minutes)
Total API requests: {stats.get('total_requests', 0)}
Requests per second: {stats.get('requests_per_second', 0):.2f}

DATA QUALITY NOTES
==================
Analysis focuses on references cited by the source articles
Combined data from Crossref and OpenAlex improves completeness
All standard statistical analyses performed (authors, journals, countries, etc.)
Error handling ensures report generation even with partial data
Affiliations normalized and grouped for consistent organization names
"""

            # Create Report_Summary tab first
            ws_summary = wb.create_sheet('Report_Summary')
            for line in summary_content.split('\n'):
                ws_summary.append([line])

            # Remove authors_surnames column and add author_count for main tables
            if not source_articles_df.empty and 'authors_surnames' in source_articles_df.columns:
                source_articles_df = source_articles_df.drop(columns=['authors_surnames'])
            if not combined_df.empty and 'authors_surnames' in combined_df.columns:
                combined_df = combined_df.drop(columns=['authors_surnames'])
            if not unique_df.empty and 'authors_surnames' in unique_df.columns:
                unique_df = unique_df.drop(columns=['authors_surnames'])
            if not duplicate_df.empty and 'authors_surnames' in duplicate_df.columns:
                duplicate_df = duplicate_df.drop(columns=['authors_surnames'])

            sheets_data = [
                ('Source_Articles', source_articles_df),
                ('All_References', combined_df),
                ('All_Unique_References', unique_df),
                ('Duplicate_References', duplicate_df),
                ('Failed_References', failed_df)
            ]

            analysis_methods = [
                ('Author_Frequency', self.analyze_authors_frequency),
                ('Journal_Frequency', self.analyze_journals_frequency),
                ('Affiliation_Frequency', self.analyze_affiliations_frequency),
                ('Country_Frequency', self.analyze_countries_frequency),
                ('Year_Distribution', self.analyze_year_distribution),
                ('5_Years_Period', self.analyze_five_year_periods)
            ]

            for sheet_name, method in analysis_methods:
                try:
                    result_df = method(combined_df)
                    sheets_data.append((sheet_name, result_df))
                except Exception as e:
                    sheets_data.append((sheet_name, pd.DataFrame()))

            try:
                title_word_data = []
                for i, (word, count) in enumerate(content_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Content_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(compound_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Compound_Words', 'Rank': i, 'Word': word, 'Frequency': count})
                for i, (word, count) in enumerate(scientific_freq.most_common(50), 1):
                    title_word_data.append({'Category': 'Scientific_Stopwords', 'Rank': i, 'Word': word, 'Frequency': count})

                title_word_df = pd.DataFrame(title_word_data)
                sheets_data.append(('Title_Word_Frequency', title_word_df))
            except Exception as e:
                sheets_data.append(('Title_Word_Frequency', pd.DataFrame()))

            for sheet_name, df in sheets_data:
                try:
                    if not df.empty:
                        ws = wb.create_sheet(sheet_name)
                        for r in dataframe_to_rows(df, index=False, header=True):
                            ws.append(r)
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([f"No data available for {sheet_name}"])
                except Exception as e:
                    pass

            wb.save(excel_path)

            return excel_path

        except Exception as e:
            try:
                timestamp = int(time.time())
                excel_path = os.path.join(tempfile.gettempdir(), f"minimal_references_analysis_results_{timestamp}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Error_Report_References"
                ws.append(["ERROR REPORT - REFERENCES ANALYSIS"])
                ws.append([f"Critical error during references analysis: {str(e)}"])
                ws.append([f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws.append(["DOIs processed:", ', '.join(doi_list)])
                wb.save(excel_path)
                return excel_path
            except:
                return "error_creating_references_report"

# =============================================
# STREAMLIT INTERFACE
# =============================================

def main():
    st.set_page_config(
        page_title="Citation Analyzer",
        page_icon="📚",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("📚 Advanced Citation Analyzer")
    st.markdown("---")
    
    # Initialize analyzer with cache
    if 'analyzer' not in st.session_state:
        st.session_state.analyzer = CitationAnalyzer()
    
    # Cache management
    if st.sidebar.button("Clear Cache"):
        cache_manager.clear_cache()
        st.session_state.analyzer = CitationAnalyzer()
        st.sidebar.success("Cache cleared!")
    
    # Main tabs
    tab1, tab2 = st.tabs(["References Analysis", "Citing Articles Analysis"])
    
    with tab1:
        st.header("References Analysis")
        st.write("Analyze the references cited by the input articles")
        
        doi_input_references = st.text_area(
            "Enter DOIs for references analysis",
            value="10.1038/s41586-023-06924-6",
            placeholder="Enter DOIs (e.g., 10.1010/XYZ, doi:10.1010/XYZ, https://doi.org/10.1010/XYZ, etc.) separated by any punctuation or newlines",
            height=150
        )
        
        if st.button("Analyze References", type="primary"):
            if doi_input_references:
                with st.spinner("Parsing DOIs..."):
                    doi_list = st.session_state.analyzer.parse_doi_input(doi_input_references)
                
                if doi_list:
                    st.success(f"Found {len(doi_list)} valid DOI(s)")
                    
                    # Main progress bar
                    main_progress = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        status_text.text("Starting sequential processing for references analysis...")
                        combined_references_df, source_articles_df, total_references, unique_dois, all_titles = st.session_state.analyzer.process_doi_sequential(doi_list)
                        
                        main_progress.progress(1.0)
                        status_text.text("Analysis complete! Generating report...")
                        
                        # Display results
                        st.subheader("Analysis Results")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total References", total_references)
                        with col2:
                            st.metric("Unique DOIs", unique_dois)
                        with col3:
                            st.metric("Processed References", len(combined_references_df) if not combined_references_df.empty else 0)
                        
                        # Show source articles
                        if not source_articles_df.empty:
                            st.subheader("Source Articles")
                            st.dataframe(source_articles_df.head(10))
                        
                        # Show unique references
                        unique_df = st.session_state.analyzer.get_unique_references(combined_references_df)
                        if not unique_df.empty:
                            st.subheader("Unique References (Sample)")
                            st.dataframe(unique_df.head(10))
                        
                        # Generate and download Excel report
                        excel_path = st.session_state.analyzer.save_all_data_to_excel(
                            combined_references_df, source_articles_df, doi_list, 
                            total_references, unique_dois, all_titles
                        )
                        
                        if os.path.exists(excel_path):
                            with open(excel_path, "rb") as f:
                                excel_data = f.read()
                            
                            st.download_button(
                                label="Download Complete Analysis Report",
                                data=excel_data,
                                file_name=os.path.basename(excel_path),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            # Clean up
                            os.remove(excel_path)
                        
                        status_text.empty()
                        
                    except Exception as e:
                        st.error(f"Error during processing: {e}")
                        main_progress.empty()
                        status_text.empty()
                else:
                    st.error("No valid DOIs found. Please check your input.")
            else:
                st.warning("Please enter at least one DOI")
    
    with tab2:
        st.header("Citing Articles Analysis")
        st.write("Find articles that cite the input articles (forward citations)")
        
        doi_input_citing = st.text_area(
            "Enter DOIs for citing articles analysis",
            value="10.1038/s41586-023-06924-6",
            placeholder="Enter DOIs (e.g., 10.1038/s41586-023-06924-6) separated by any punctuation or newlines",
            height=150
        )
        
        if st.button("Analyze Citing Articles", type="secondary"):
            if doi_input_citing:
                with st.spinner("Parsing DOIs..."):
                    doi_list = st.session_state.analyzer.parse_doi_input(doi_input_citing)
                
                if doi_list:
                    st.success(f"Found {len(doi_list)} valid DOI(s)")
                    
                    # Main progress bar
                    main_progress = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        status_text.text("Starting sequential processing for citing articles analysis...")
                        citing_articles_df, citing_details_df, citing_results, all_citing_titles = st.session_state.analyzer.process_citing_articles_sequential(doi_list)
                        
                        main_progress.progress(1.0)
                        status_text.text("Analysis complete! Generating report...")
                        
                        # Display results
                        st.subheader("Citing Articles Analysis Results")
                        
                        if citing_results:
                            total_citation_relationships = len(citing_articles_df) if citing_articles_df is not None else 0
                            total_unique_citations = len(st.session_state.analyzer.get_unique_citations(citing_articles_df)) if citing_articles_df is not None and not citing_articles_df.empty else 0
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Source Articles", len(doi_list))
                            with col2:
                                st.metric("Citation Relationships", total_citation_relationships)
                            with col3:
                                st.metric("Unique Citing Articles", total_unique_citations)
                            
                            st.subheader("Citations per Source Article")
                            for doi, data in citing_results.items():
                                st.write(f"- **{doi}**: {data['count']} citations")
                            
                            if citing_articles_df is not None and not citing_articles_df.empty:
                                st.subheader("Citing Articles (Sample)")
                                display_cols = ['source_doi', 'doi', 'title', 'authors_with_initials', 'author_count', 'year', 'journal_abbreviation',
                                              'citation_count_openalex']
                                st.dataframe(citing_articles_df[display_cols].head(10))
                            
                            # Generate and download Excel report
                            excel_path = st.session_state.analyzer.save_citation_analysis_to_excel(
                                citing_articles_df, citing_details_df, doi_list, citing_results, all_citing_titles
                            )
                            
                            if os.path.exists(excel_path):
                                with open(excel_path, "rb") as f:
                                    excel_data = f.read()
                                
                                st.download_button(
                                    label="Download Citation Analysis Report",
                                    data=excel_data,
                                    file_name=os.path.basename(excel_path),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                
                                # Clean up
                                os.remove(excel_path)
                        else:
                            st.info("No citing articles found for the provided DOIs.")
                        
                        status_text.empty()
                        
                    except Exception as e:
                        st.error(f"Error during processing: {e}")
                        main_progress.empty()
                        status_text.empty()
                else:
                    st.error("No valid DOIs found. Please check your input.")
            else:
                st.warning("Please enter at least one DOI")

if __name__ == "__main__":
    main()



